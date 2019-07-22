#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib
import urllib2
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
sys.path.append(".")
sys.path.append("..")
from internal.ts_common import *

#添加trade data，在满足条件的行中

dpath = "../data/"

# Main
pindex = len(sys.argv)
if pindex<2:
	sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " 代码\n")
	exit(1);

code = sys.argv[1]
if (len(code) != 6):
	sys.stderr.write("Len should be 6\n")
	exit(1);

ret, code = parseCode(code)
if ret!=0:
	exit(1);

prepath = dpath + code + "/"
items = os.listdir(prepath)
items.reverse()

for f in items:
	filexlsx = prepath + f
	if(os.path.isdir(filexlsx)):
		continue
	date = f[9:19]
	list = []
	list.append(date)
	#print f,date
	wb = load_workbook(filexlsx)
	sheetnames = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheetnames[0])
	#ws = wb.active

	#print "Work Sheet Titile:", ws.title
	#print "Work Sheet Rows:", ws.max_row
	#print "Work Sheet Cols:", ws.max_column
	update_row_data(ws, f)
	wb.save(filexlsx)
print "Update row Fin"
