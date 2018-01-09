#!/usr/bin/env python
# -*- coding:gbk -*-
#可转债
import sys
import re
import urllib2
import datetime
import zlib
import json
import getopt
import pandas as pd
sys.path.append(".")
sys.path.append("..")
from internal.dfcf_interface import *
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

#根据参数不同操作
#1.没有参数，访问DC将数据显示在控制台
#2.多个参数 sdtae
#

'''
urlall = "http://data.eastmoney.com/kzz/default.html"

http://dcfm.eastmoney.com/em_mutisvcexpandinterface/api/js/get?type=KZZ_LB&token=70f12f2f4f091e459a279469fe49eca5&cmd=&st=STARTDATE&sr=-1&p=2&ps=50&js=var%20fnhOGnNf={pages:(tp),data:(x)}&rt=50462432
curl 'http://dcfm.eastmoney.com/em_mutisvcexpandinterface/api/js/get?type=KZZ_LB&token=70f12f2f4f091e459a279469fe49eca5&cmd=&st=STARTDATE&sr=-1&p=2&ps=50&js=var%20ShTNyQGa=\{pages:(tp),data:(x)\}&rt=50462435' -H 'Referer: http://data.eastmoney.com/kzz/default.html' -H 'User-Agent: Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2986.0 Safari/537.36' --compressed ;
curl 'http://datapic.eastmoney.com/img/loading.gif' -H 'Referer: http://data.eastmoney.com/kzz/default.html' -H 'User-Agent: Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2986.0 Safari/537.36' --compressed
==========
[STARTDATE]='2017-12-04T00:00:00'		申购日期
[QSCFJ]='7.31'				强赎触发价
[TEXCH]='CNSESZ'			交易所代码
[AISSUEVOL]='10.0'			发行规模
[ISSUEPRICE]='100.0'		发行价
[DELISTDATE]='-'			摘牌日期
[BCODE]='17270600001JBV'	？？编码
[SWAPSCODE]='002284'		正股代码
[CORRESNAME]='亚太发债'		债券发行名称
[BONDCODE]='128023'			债券代码
[PB]='2.36'					PB
[SNAME]='亚太转债'			债券名称
[PARVALUE]='100.0'			？？
[MarketType]='2'			市场类型 1：上交，2：深交
[LISTDATE]='2017-12-26T00:00:00'		上市交易日期
[HSCFJ]='13.57'				回售触发价
[SECURITYSHORTNAME]='亚太股份'			正股名称
[GDYX_STARTDATE]='2017-12-01T00:00:00'	股权登记日
[LUCKRATE]='0.0245376003'	中签率
[ZGJZGJJZ]='81.6091954023'	当前转股价值
[LIMITBUYIPUB]='100.0'		最大申购上限
[MEMO]=''					备注
[ZGJ]='8.52'				转股价
[ZQHDATE]='2017-12-06T00:00:00'			中签号发布日
[YJL]='22.5352112676'		溢价率
[FSTPLACVALPERSTK]='1.3558'	每股配售额
[ZQNEW]='100.0'				债券最新价格
[CORRESCODE]='072284'		申购代码
[SWAPPRICE]='10.44'			？？初始转股价
[ZGJ_HQ]='8.52'				正股价
[ZGJZGJ]='10.44'			当前转股价
'''

urlfmt = 'http://dcfm.eastmoney.com/em_mutisvcexpandinterface/api/js/get?type=KZZ_LB&token=70f12f2f4f091e459a279469fe49eca5&cmd=&st=STARTDATE&sr=-1&p=%d&ps=50&js=iGeILSKk={pages:(tp),data:(x)}&rt=50463927'
send_headers = {
'Host':'dcfm.eastmoney.com',
'Connection':'keep-alive',
'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.101 Safari/537.36',
'Accept':'*/*',
'DNT': '1',
'Referer': 'http://data.eastmoney.com/kzz/default.html',
'Accept-Encoding': 'gzip, deflate',
'Accept-Language': 'zh-CN,zh;q=0.8',
'Cookie': 'st_pvi=33604263703666; emstat_bc_emcount=300285734935048278; \
pi=6100112247957528%3byudingding6197%3bgoutou%3bcUC1rQLS6GFOJIpJ%2b0I6Wt5AdIat%2bLRj2ZvGrlzeObRNvIHEcy62FDfQ%2boIImkvxiIyCd9QIklChsWI2qjINWL\
5DdBKYMZ71JGBsgoVjEXYjdw1rWDHu45I%2bNugmP4pbtdgvhUf884FcXhI1tqTCeHOobtdLSzpfA7h3MiSCx5rf8AdOH0uOhUuvYFxLUOx0oD6KGdMI%3bJ7wwpyq2YPDBfbw\
AqENYGA8GKWnFXYc1dIR5LuUNwKNYfZKtn2ufQSBXaOy%2fJuok5A10Hmp70CM%2bH4jRSUeLe8OOEOwSG5o1tvO4rx%2fTjNoq%2fbM2d6QYkUKtXL0XTX8nREubTh%2bPugi\
WdGxX3hARJpanE0qULw%3d%3d; \
uidal=6100112247957528goutou; vtpst=|; '
}

rturl = 'http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeDataSimple?page=1&num=80&sort=symbol&asc=1&node=hskzz_z&_s_r_a=page'
rt_headers = {
'Host': 'vip.stock.finance.sina.com.cn',
'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.101 Safari/537.36',
'Content-type': 'application/x-www-form-urlencoded',
'Accept': '*/*',
'DNT': 1,
'Referer': 'http://vip.stock.finance.sina.com.cn/mkt/',
'Accept-Encoding': 'gzip, deflate',
'Accept-Language': 'zh-CN,zh;q=0.8',
'Cookie': 'U_TRS1=0000000e.db0f67d9.58743532.0ab3855e; UOR=www.baidu.com,blog.sina.com.cn,; vjuids=-4f7e5f1b8.15985efc8ab.0.b0009b04a9d3a; \
SINAGLOBAL=114.243.223.213_1484010803.467733; SGUID=1490330513641_6143460; \
SCF=ApQZBkYNx5ED9eRh4x7RWZjDJZfZGEsCEcgqoaFHnaP7DqJZQpUkYRbUtwv1spWbrMvv9eU5YBJ8U5RXwjUggcc.; \
FINA_V_S_2=sz300648,sh603600; vjlast=1504425305; Apache=10.13.240.35_1513310793.422171; U_TRS2=00000016.b8c612f7.5a3398c2.608c14c5; \
SessionID=oe6kbh7j9v4usqdkigqe4inb71; ULV=1513330875049:56:5:3:10.13.240.35_1513310793.422171:1513225944670; \
sso_info=v02m6alo5qztbmdlpGpm6adpJqWuaeNo4S5jbKZtZqWkL2Mk5i1jaOktYyDmLOMsMDA; \
SUB=_2A253Rcj3DeThGedI7lQY9S7KyD-IHXVUMr0_rDV_PUNbm9AKLWTjkW9NVwM9cn_D0fMlGi8-URaLNK3j_mTGomwb; \
SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5.n90RO.U79QHoy5Rk17Op5NHD95QpSo-c1K-7Soe0Ws4Dqcjdi--NiKyFiK.Ni--4i-zpi-ihi--fi-2Xi-zX; \
ALF=1545792551; rotatecount=2; SR_SEL=1_511; lxlrttp=1513341002; FINANCE2=8d5626b3132364178ca15d9e87dc4f27; \
SINA_FINANCE=yudingding6197%3A1656950633%3A4'
}

'''
def getKZZConnect(page):
	global content
	LOOP_COUNT=0
	urllink = urlfmt % (page)
	res_data = None
	while LOOP_COUNT<3:
		try:
			#print urllink
			req = urllib2.Request(urllink,headers=send_headers)
			res_data = urllib2.urlopen(req)
		except:
			print "Exception kzz urlopen"
			LOOP_COUNT += 1
		else:
			break
	if res_data is None:
		print "Error: Fail to get request"
		content = ''
		return
	content = res_data.read().decode('utf8')
	return
'''

#得到交易日KZZ的实时价格
def getKZZRtSina(list):
	global content
	LOOP_COUNT=0
	urllink = rturl
	res_data = None
	while LOOP_COUNT<3:
		try:
			#print urllink
			req = urllib2.Request(urllink,headers=rt_headers)
			res_data = urllib2.urlopen(req)
		except:
			print "Exception kzz urlopen"
			LOOP_COUNT += 1
		else:
			break
	#print res_data
	if res_data is None:
		print "Error: Fail to get request"
		content = ''
		return
	content = res_data.read()
	respInfo = res_data.info()
	if( ("Content-Encoding" in respInfo) and (respInfo['Content-Encoding'] == "gzip")) :
		content = zlib.decompress(content, 16+zlib.MAX_WBITS);
	#print content
	if len(content)<=2:
		content = ''
		return
	if content[0]=='[' and content[-1]==']':
		content = content [1:-1]

	# 'settlement'是昨收
	conv_clmn=['changepercent','trade','pricechange','open','high','low','settlement']
	while 1:
		obj = re.match(r'{(.*?)},?(.*)', content)
		if obj is None:
			break
		record = obj.group(1)
		#print record
		content = obj.group(2)
		dict = {}
		ll = record.split(',')
		for item in ll:
			#第二项表示非"字符进行匹配
			dictObj = re.match(r'(.*?):"?([^"]*)', item)
			if dictObj is None:
				print "Invalid item in dict", item
				break
			#将symbol前面的sh or sz 去掉
			if dictObj.group(1)=='symbol':
				dict[dictObj.group(1)] = dictObj.group(2)[2:]
			#将涨跌百分比强转float
			elif dictObj.group(1) in conv_clmn:
				dict[dictObj.group(1)] = float(dictObj.group(2))
			else:
				dict[dictObj.group(1)] = dictObj.group(2)
		list.append(dict)
	return

#['BONDCODE','SNAME','ZQNEW','YJL','ZGJZGJJZ','ZGJ_HQ','SWAPSCODE','SECURITYSHORTNAME','ZGJZGJ']
# 债券最新价格,溢价率,当前转股价值,正股价??,当前转股价
def output(kdf):
	fmt = "%2d %6s %8s	%8s   %5.2f	%5.2f	%5.2f	%5.2f"
	df = kdf.sort_values(['YJL'],0,True)
	rank = 0
	for index,items in df.iterrows():
		value=items[1]
		nmlen=len(items[1].encode('gbk'))
		if nmlen<8:
			left=8-nmlen
			while left>0:
				value=' '+value
				left-=1

		ZQNEW = items['ZQNEW']
		YJL = float(items['YJL'])
		ZGJZ = float(items['ZGJZGJJZ'])
		if items['ZGJ_HQ']=='-':
			ZGJ_HQ = 0
		else:
			ZGJ_HQ = float(items['ZGJ_HQ'])
		ZGJZGJ = float(items['ZGJZGJ'])

		str = fmt % (rank+1, items[0],items[1],ZQNEW,YJL,ZGJZ,ZGJ_HQ,ZGJZGJ)
		rank += 1
		print str
	return

def saveToExcel(kzzList):
	if len(kzzList)==0:
		return
	
	wb = Workbook()
	ws = wb.active
	strline = u'code,名称,正股价格,转股价,转股价值,债价格,溢价率,n股价,n转股价值,n溢价率'
	strObj = strline.split(u',')
	ws.append(strObj)
	#随着列数进行改变
	ws.auto_filter.ref = "A1:J1"
	excel_row = 2

	k = 0
	ascid = 65
	for item in kzzlist:
		itemList = []
		itemList.append(item['BONDCODE'])
		itemList.append(item['SNAME'])
		itemList.append(item['ZGJ_HQ'])
		itemList.append(item['ZGJZGJ'])
		ZGJZGJJZ_p = round(float(item['ZGJZGJJZ']), 2)
		itemList.append(ZGJZGJJZ_p)
		itemList.append(item['ZQNEW'])
		YJL_p = round(float(item['YJL']), 2)
		itemList.append(YJL_p)

		itemList.append(item['ZGJ_HQ'])
		itemList.append(ZGJZGJJZ_p)
		itemList.append(YJL_p)

		number = len(itemList)
		for k in range(0,number):
			cell = chr(ascid+k) + str(excel_row)
			ws[cell] = itemList[k]
		excel_row += 1

	folder = '../data/entry/kzz'
	if not os.path.exists(folder):
		os.makedirs(folder)
	today = datetime.date.today()
	tdstr = '%04d-%02d-%02d' %(today.year, today.month, today.day)
	filexlsx1 = '%s/kzzInfo_%s.xlsx' %(folder, tdstr)
	wb.save(filexlsx1)
	print "Save file", filexlsx1
	return

param_config = {
	"Daoxu":0,
	"NoDetail":0,
	"SortByTime":0,
	"AllInfo":0,
	"Excel":0
}

#Main
if __name__=="__main__":
	optlist, args = getopt.getopt(sys.argv[1:], '?sdtae')
	for option, value in optlist:
		if option in ["-s","--sort"]:
			param_config["Daoxu"] = 1
		elif option in ["-d","--nodetail"]:
			param_config["NoDetail"] = 1
		elif option in ["-t","--sbtime"]:
			param_config["SortByTime"] = 1
		elif option in ["-a","--all"]:
			param_config["AllInfo"] = 1
		elif option in ["-e","--excel"]:
			param_config["Excel"] = 1
		elif option in ["-?","--???"]:
			print "Usage:", os.path.basename(sys.argv[0]), " [-sdtae]"
			print "-s: sort DaoXu"
			print "-d: NoDetail"
			print "-t: sort by time"
			print "-a: All Info, building..."
			print "-e: save as Excel"
			exit()

	req_count=0
	curpage = 1
	kzzlist = []
	content=''
	totalpage = 0
	while 1:
		req_count += 1
		if req_count>10:
			break

		content = getKZZConnect(urlfmt, send_headers, curpage)
		if content=='':
			break
		dataObj = re.match('^iGeILSKk={pages:(\d+),data:\[(.*)\]}', content)
		if dataObj is None:
			print "Content format not match"
			break
		if totalpage < 1:
			totalpage = int(dataObj.group(1))
		if curpage>totalpage:
			break
		curpage += 1
		#分析具体的每一项数据
		if param_config['Excel']==1:
			getFilterZhai(dataObj.group(2), 2, kzzlist)
		else:
			getFilterZhai(dataObj.group(2), 1, kzzlist)
		#print curpage, totalpage
		#break

	if len(kzzlist)==0:
		print "Not find data"
		exit()

	if param_config['Excel']==1:
		saveToExcel(kzzlist)
		exit()

	keylist = ['BONDCODE','SNAME','ZQNEW','YJL','ZGJZGJJZ','ZGJ_HQ','SWAPSCODE','SECURITYSHORTNAME','ZGJZGJ']
	kzzdf = pd.DataFrame(kzzlist, columns=keylist)
	output(kzzdf)


