# -*- coding:gbk -*-
import sys
import re
import os
import string
import urllib
import urllib2
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

fm_url = "http://datainterface.eastmoney.com/EM_DataCenter/JS.aspx?type=NS&sty=NSSTV5&st=12&sr=-1&p=%d&ps=50"
prepath1 = "..\\Data\\entry\\xingu\\"

def handle_stk_data(stk_item, stockInfo):
	str_arr = stk_item.split(',')
	name = str_arr[3]
	code = str_arr[4]
	daxin_code = str_arr[5]
#	if code=='300618':
#		for i in range(0, len(str_arr)):
#			print "%02d	%s"%(i, str_arr[i])
	#print "%6s	%s	%s"%(code, str_arr[31], str_arr[34])

	intro = str_arr[27]
	faxing_liang = int(float(str_arr[6]))
	wangshang_fx = int(float(str_arr[7]))
	print code,name,faxing_liang
	faxing_jia=str_arr[10]
	if faxing_jia!='':
		faxing_jia = float(faxing_jia)
	shengou_date = str_arr[11]
	zhongqian_date = str_arr[12]
	shangshi_date = str_arr[13]
	zhongqian_lv = str_arr[15]
	if zhongqian_lv!='':
		zhongqian_lv = float(zhongqian_lv)
	xunjia_mul = str_arr[17]
	ipo_close = str_arr[24]
	if ipo_close!='':
		ipo_close = float(ipo_close)
	#行业市盈率
	hangye_shiying = str_arr[32]
	if hangye_shiying!='':
		hangye_shiying = float(hangye_shiying)
	#配售对象报价家数
	baojia_count = str_arr[33]
	if baojia_count!='':
		baojia_count = float(baojia_count)
	#发行状态
	faxing_state = str_arr[38]
	if faxing_state=='':
		faxing_state = '开板'.decode('gbk')
	
	stockInfo.append(code)
	stockInfo.append(name)
	stockInfo.append(daxin_code)
	stockInfo.append(faxing_liang)
	stockInfo.append(faxing_jia)
	stockInfo.append(faxing_state)
	stockInfo.append(wangshang_fx)
	stockInfo.append(shengou_date)
	stockInfo.append(zhongqian_date)
	stockInfo.append(zhongqian_lv)
	stockInfo.append(shangshi_date)
	stockInfo.append(xunjia_mul)
	stockInfo.append(ipo_close)
	stockInfo.append(hangye_shiying)
	stockInfo.append(baojia_count)
	stockInfo.append(intro)

wb = Workbook()
# grab the active worksheet
ws = wb.active
strline = u'code,名称,申购代码,发行量,发行价,发行状态,网上发行量,申购日期,中签公布日,中签率,上市日期,询价累计数,首日收盘价,行业市盈率,报价家数,公司介绍'
strObj = strline.split(u',')
ws.append(strObj)
#随着列数进行改变
ws.auto_filter.ref = "A1:P1"
excel_row = 2
repeat_flag = 0
first_code = ''

for i in range(1, 35):
	if repeat_flag==1:
		break

	url = fm_url%(i)
	#print url
	#response type is 'instance'
	response = None
	LOOP_COUNT = 0
	while LOOP_COUNT<3:
		try:
			req = urllib2.Request(url)
			response = urllib2.urlopen(req, timeout=5)
		except:
			LOOP_COUNT += 1
			print "URL request timeout"
		else:
			break
	if response is None:
		break

	line = response.readline()
	while line:
		line = line.decode('utf8')

		obj = re.match('.*\(\[(.*)\]\)',line)
		line = obj.group(1)
		#这个if可能判断多余，强行左右匹配引号里的内容""
		if line is None:
			pos = line.find('"')
			if pos == -1:
				line = response.readline()
				continue
			rpos = line.rfind('"')
			if rpos==-1:
				line = response.readline()
				continue
			if rpos<=pos:
				line = response.readline()
				continue
			line = line[pos:rpos+1]
		firstln = 1
		while 1:
			obj = re.match(r'"(.*?)",?(.*)', line)
			if obj is None:
				break
			line = obj.group(2)

			stockInfo = []
			stk_item = obj.group(1)
			
			if firstln==1:
				str_arr = stk_item.split(',')
				code = str_arr[4]
				#print "%d '%s'=='%s' "%(i,first_code,code)
				if first_code==code:
					repeat_flag = 1
					break
				first_code = code
				firstln = 0

			handle_stk_data(stk_item, stockInfo)
			#添加到表格中
			k = 0
			ascid = 65
			number = len(stockInfo)
			for k in range(0,number):
				cell = chr(ascid+k) + str(excel_row)
				ws[cell] = stockInfo[k]
			excel_row += 1
		line = response.readline()

today = datetime.date.today()
cur=datetime.datetime.now()
qdate = '%04d-%02d-%02d' %(today.year, today.month, today.day)
filexlsx1 = prepath1 + "xg_fx"+ qdate
filexlsx1 = '%s_%02d-%02d.xlsx' %(filexlsx1, cur.hour, cur.minute)
wb.save(filexlsx1)

file2 = prepath1 + "last_xg.xlsx"
shutil.copy(filexlsx1, file2)
