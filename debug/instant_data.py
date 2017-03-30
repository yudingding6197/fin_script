# -*- coding:gbk -*-
import sys
import re
import os
import string
import urllib
import urllib2
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

#获取每天的所有的信息，包含trade和停牌的
#放到 "..\\Data\\entry\\trade\\trade_****.xlxs"

#'code,名称,涨跌幅,收盘价,涨跌价,振幅,成交量,成交额,昨收,今开,最高,最低,5分钟幅,Item1,Item2,Item3'

def handle_float(str):
	if str=='-':
		obj = None
	else:
		obj = float(str)
	return obj

def handle_int(str):
	if str=='-':
		obj = None
	else:
		obj = int(str)
	return obj

def handle_stk_data(stk_item, stockInfo):
	str_arr = stk_item.split(',')
	name = str_arr[2]
	code = str_arr[1]
	#if code=='603999':
	#	for i in range(0, len(str_arr)):
	#		print "%02d	%s"%(i, str_arr[i])
	#print "%6s	%s	%s"%(code, str_arr[31], str_arr[34])

	close = str_arr[3]
	close = handle_float(close)

	change_price = str_arr[4]
	change_price = handle_float(change_price)

	change_perc = str_arr[5]
	if change_perc=='-':
		change_perc = None
	else:
		change_perc = change_perc[:-1]
		change_perc = float(change_perc)

	zhengfu = str_arr[6]
	zhengfu = handle_float(zhengfu)

	volume = str_arr[7]
	volume = handle_int(volume)

	amount = str_arr[8]
	if amount=='-':
		amount = None
	else:
		amount = long(amount)

	pre_close = str_arr[9]
	if pre_close=='0.00':
		pre_close = None
	else:
		pre_close = float(pre_close)

	open = str_arr[10]
	open = handle_float(open)
	high = str_arr[11]
	high = handle_float(high)
	low = str_arr[12]
	low = handle_float(low)
	#5分钟涨幅
	zf_5min = str_arr[21]
	if zf_5min=='-':
		zf_5min = None
	else:
		zf_5min = zf_5min[:-1]
		zf_5min = float(zf_5min)
	item1 = str_arr[22]
	item1 = handle_float(item1)
	item2 = str_arr[23]
	item2 = handle_float(item2)
	item3 = str_arr[24]
	item3 = handle_float(item3)
	
	stockInfo.append(code)
	stockInfo.append(name)
	stockInfo.append(change_perc)
	stockInfo.append(close)
	stockInfo.append(change_price)
	stockInfo.append(zhengfu)
	stockInfo.append(volume)
	stockInfo.append(amount)
	stockInfo.append(pre_close)
	stockInfo.append(open)
	stockInfo.append(high)
	stockInfo.append(low)
	stockInfo.append(zf_5min)
	stockInfo.append(item1)
	stockInfo.append(item2)
	stockInfo.append(item3)

# Main
fm_url = "http://nufm.dfcfw.com/EM_Finance2014NumericApplication/JS.aspx?type=CT&cmd=C._A&sty=FCOIATA&sortType=A&sortRule=-1&page=%d&pageSize=%d&token=7bc05d0d4c3c22ef9fca8c2a912d779c"
prepath1 = "..\\Data\\entry\\trade\\"
name = "trade"
PAGE_COUNT = 500

wb = Workbook()
# grab the active worksheet
ws = wb.active
strline = u'code,名称,涨跌幅,收盘价,涨跌价,振幅,成交量,成交额,昨收,今开,最高,最低,5分钟幅,Item1,Item2,Item3'
strObj = strline.split(u',')
ws.append(strObj)
#随着列数进行改变
ws.auto_filter.ref = "A1:P1"
excel_row = 2
repeat_flag = 0
first_code = ''

for i in range(1, 35):
	if repeat_flag==1:
		break

	url = fm_url%(i,PAGE_COUNT)
	#print url
	#response type is 'instance'
	response = None
	LOOP_COUNT = 0
	while LOOP_COUNT<3:
		try:
			req = urllib2.Request(url)
			response = urllib2.urlopen(req, timeout=5)
		except:
			LOOP_COUNT += 1
			print "URL request timeout"
		else:
			break
	if response is None:
		break

	line = response.readline()
	while line:
		line = line.decode('utf8')

		obj = re.match('.*\(\[(.*)\]\)',line)
		line = obj.group(1)
		#这个if可能判断多余，强行左右匹配引号里的内容""
		if line is None:
			pos = line.find('"')
			if pos == -1:
				line = response.readline()
				continue
			rpos = line.rfind('"')
			if rpos==-1:
				line = response.readline()
				continue
			if rpos<=pos:
				line = response.readline()
				continue
			line = line[pos:rpos+1]

		total_item = 0
		#记录每一页第一个代码，避免重复读取数据
		firstln = 1
		while 1:
			obj = re.match(r'"(.*?)",?(.*)', line)
			if obj is None:
				break
			line = obj.group(2)

			stockInfo = []
			stk_item = obj.group(1)
			
			if firstln==1:
				str_arr = stk_item.split(',')
				code = str_arr[1]
				#print "%d '%s'=='%s' "%(i,first_code,code)
				if first_code==code:
					repeat_flag = 1
					break
				first_code = code
				firstln = 0

			handle_stk_data(stk_item, stockInfo)
			#添加到表格中
			k = 0
			ascid = 65
			number = len(stockInfo)
			for k in range(0,number):
				cell = chr(ascid+k) + str(excel_row)
				ws[cell] = stockInfo[k]
			excel_row += 1
			total_item += 1
		#假设最后一页取出数据太少，停止http请求
		if total_item<PAGE_COUNT-10:
			break
		line = response.readline()

today = datetime.date.today()
cur=datetime.datetime.now()
qdate = '%04d-%02d-%02d' %(today.year, today.month, today.day)
filexlsx1 = prepath1 + name + qdate
filexlsx1 = '%s#%02d-%02d.xlsx' %(filexlsx1, cur.hour, cur.minute)
wb.save(filexlsx1)
