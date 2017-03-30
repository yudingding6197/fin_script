# -*- coding:gbk -*-
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
# 中文

# Main
pindex = len(sys.argv)
if pindex<2:
	sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " 代码\n")
	exit(0)
code = sys.argv[1]

path = "..\\buy_sell\\"
file_list = []
filter_item = "A1:I1"
for f in os.listdir(path):
	if os.path.isfile(path + f) is False:
		continue
	file_list.append(f)

c_list = ['date','time','code','name','op','vol','price','amount']
df = pd.DataFrame()
st_date = file_list[0][6:12]
ed_date = file_list[-1][6:12]
print st_data,ed_date
for file in file_list:
	dt_str = file[6:12]
	if dt_str.isdigit() is False:
		print "Invalid file(%s) or date(%s)" % (file, dt_str)
		continue

	sheet_st = 'table'
	wb = load_workbook(path+file)
	ws = wb.get_sheet_by_name(sheet_st)
	for rx in range(2, ws.max_row+1):
		w2 = "%06d" % (w2)
		if w2!=code:
			continue

		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value
		w7 = ws.cell(row = rx, column = 7).value

		temp_list = [int(dt_str),w1,w2,w3,w4,w5,w6,w7]
		df1 = pd.DataFrame([temp_list], columns=c_list)
		df = df.append(df1)
		#print temp_list

if len(df)>0:
	filename = "%s%s%s%d_%d.xlsx" %(path, "trade\\", "statics_", st_date, ed_date)
	df.to_excel(filename)
