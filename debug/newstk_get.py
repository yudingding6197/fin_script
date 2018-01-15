#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import string
import urllib
import urllib2
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
sys.path.append(".")
sys.path.append("..")
from internal.ts_common import *

#��ȡÿ������е���Ϣ����new stockѡ��
#�ŵ� "../data/entry/trade/ns_info.xlxs"

#'code,name,timeToMarket,outstanding,totals'

def handle_float(str):
	if str=='-':
		obj = None
	else:
		obj = float(str)
	return obj

def handle_int(str):
	if str=='-':
		obj = None
	else:
		obj = int(str)
	return obj

def check_new_stk(stk_item, stockInfo):
	str_arr = stk_item.split(',')
	name = str_arr[2]
	code = str_arr[1]
	if name[0:1]!='N':
		return 0

	url_liut = "http://vip.stock.finance.sina.com.cn/corp/go.php/vCI_StockStructureHistory/stockid/%s/stocktype/LiuTongA.phtml"
	url_totl = "http://vip.stock.finance.sina.com.cn/corp/go.php/vCI_StockStructureHistory/stockid/%s/stocktype/TotalStock.phtml"
	ltgb_list = []
	#ͨ��SINA�õ���ͨ�ɱ����ܹɱ��ı仯��Ϣ��������ϢΪһ�����߶�����ƽṹ
	#<set name='96/07' value='0' hoverText='1996-07-22'/>
	#<set name='...' value='...' hoverText='...'/>
	gb_str = get_guben_line(url_liut, code)
	parse_guben(gb_str, ltgb_list)

	zgb_list = []
	gb_str = get_guben_line(url_totl, code)
	parse_guben(gb_str, zgb_list)

	gb_len = len(ltgb_list)
	if gb_len != len(zgb_list):
		print code, name, "guben not match"
		return 0
	#�������ͣ���к����飬��Ϊ�¹����У���Ҫ���⴦��
	#���Դ����һ�п�ʼ���㣬�õ���Чֵ�󷵻�
	for i in range(0, gb_len):
		ltobj = ltgb_list[gb_len-i-1]
		zobj = zgb_list[gb_len-i-1]
		if ltobj[0] != zobj[0]:
			print code, "guben date not match", ltobj[0], zobj[0]
			return 0
		if ltobj[1]==0:
			continue

		sdate = ''.join(ltobj[0].split('-'))
		stockInfo.append(code)
		stockInfo.append(name)
		stockInfo.append(long(sdate))
		stockInfo.append(ltobj[1])
		stockInfo.append(zobj[1])
		break;
	return 1

# Main
fm_url = "http://nufm.dfcfw.com/EM_Finance2014NumericApplication/JS.aspx?type=CT&cmd=C._A&sty=FCOIATA&sortType=A&sortRule=-1&page=%d&pageSize=%d&token=7bc05d0d4c3c22ef9fca8c2a912d779c"
prepath1 = "../data/entry/trade/"
name = "ns_info.xlsx"
PAGE_COUNT = 50

wb = Workbook()
# grab the active worksheet
ws = wb.active
strline = u'code,name,timeToMarket,outstanding,totals'
strObj = strline.split(u',')
ws.append(strObj)
#�����������иı�
ws.auto_filter.ref = "A1:P1"
excel_row = 2
repeat_flag = 0
first_code = ''

#����rsp.readlineʧ�ܣ��ض���ҳ�����3��
SAME_PAGE_CT=0
i = 0
while i<100:
	if repeat_flag==1:
		break
	if SAME_PAGE_CT>3:
		break
	SAME_PAGE_CT += 1

	url = fm_url%(i,PAGE_COUNT)
	#print url
	#response type is 'instance'
	response = None
	LOOP_COUNT = 0
	while LOOP_COUNT<3:
		try:
			req = urllib2.Request(url)
			response = urllib2.urlopen(req, timeout=60)
		except:
			LOOP_COUNT += 1
			print "URL request timeout"
		else:
			break
	if response is None:
		break

	#print response
	try:
		line = response.readline()
	except:
		print "Warning read fail in page i=", i
		continue
	while line:
		line = line.decode('utf8')

		obj = re.match('.*\(\[(.*)\]\)',line)
		line = obj.group(1)
		#���if�����ж϶��࣬ǿ������ƥ�������������""
		if line is None:
			pos = line.find('"')
			if pos == -1:
				line = response.readline()
				continue
			rpos = line.rfind('"')
			if rpos==-1:
				line = response.readline()
				continue
			if rpos<=pos:
				line = response.readline()
				continue
			line = line[pos:rpos+1]

		total_item = 0
		#��¼ÿһҳ��һ�����룬�����ظ���ȡ����
		firstln = 1
		while 1:
			obj = re.match(r'"(.*?)",?(.*)', line)
			if obj is None:
				break
			line = obj.group(2)

			stockInfo = []
			stk_item = obj.group(1)
			
			if firstln==1:
				str_arr = stk_item.split(',')
				code = str_arr[1]
				if first_code==code:
					repeat_flag = 1
					break
				first_code = code
				firstln = 0

			ret = check_new_stk(stk_item, stockInfo)
			if ret==0:
				continue
			print stockInfo
			#��ӵ������
			k = 0
			ascid = 65
			number = len(stockInfo)
			for k in range(0,number):
				#FK,������26����ĸ�����ӣ�ͨ��AA1,AA2...����,�˴�����
				l = k
				dv = l/26
				md = l%26
				if dv==0:
					cell = chr(ascid+k) + str(excel_row)
				else:
					cell = chr(ascid+dv-1) + chr(ascid+md) + str(excel_row)
				ws[cell] = stockInfo[k]
			excel_row += 1
			total_item += 1
		#�������һҳȡ������̫�٣�ֹͣhttp����
		if total_item<PAGE_COUNT-5:
			break
		line = response.readline()
	i += 1
	SAME_PAGE_CT = 0

filexlsx1 = prepath1 + name
wb.save(filexlsx1)
print "DONE!"
