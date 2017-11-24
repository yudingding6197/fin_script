#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

# 读取../data/entry/cixin/下最新的cixin_***.xlsx文件，得到未打开的次新值
# 首先手动执行cixin_data.py得到最新的CX信息，再从中解析
# Main
cx_path = "../data/entry/_no_open_cx.txt"
path = "../data/entry/cixin/"
file = ''
for (dirpath, dirnames, filenames) in os.walk(path):
	if len(filenames)>0:
		file = filenames[-1]
	#print('dirpath = ' + dirpath)
if file=='':
	print "Not find file in path:", path
	exit(0)

yzcx_df = pd.DataFrame()
sheet_st = 'Sheet'
wb = load_workbook(path+file)
ws = wb.get_sheet_by_name(sheet_st)
code_list = []
for rx in range(2,ws.max_row+1):
	w1 = ws.cell(row = rx, column = 1).value
	w2 = ws.cell(row = rx, column = 2).value
	w3 = ws.cell(row = rx, column = 3).value
	w4 = ws.cell(row = rx, column = 4).value

	if int(w4)==0:
		temp_list = [w1,w2,w3,w4]
		df1 = pd.DataFrame([temp_list])
		yzcx_df = yzcx_df.append(df1)
		code_list.append(w1)

if len(code_list)==0:
	print "No CIXIN List"
	exit(0)

cx_file = open(cx_path, 'w')
for code in code_list:
	cx_file.write(code + "\n")
cx_file.close()
