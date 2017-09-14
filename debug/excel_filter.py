#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

#���ļ����岻�󣬽�../Data/s****/��һ��xlsx�ļ�A1-G1���Filter
prepath = '../Data/'
allFileNum = 0
volumnList = [200, 300, 600, 900]
sheet_st = "Sheet"

def printPath(level, path,fileList):
	global allFileNum
	# �����ļ��У���һ���ֶ��Ǵ�Ŀ¼�ļ���
	dirList = []
	# �����ļ�
	#fileList = []
	# ����һ���б����а�����Ŀ¼��Ŀ������(google����)
	files = os.listdir(path)
	# �����Ŀ¼����
	dirList.append(str(level))
	for f in files:
		if(os.path.isdir(path + '/' + f)):
			# �ų������ļ��С���Ϊ�����ļ��й���
			if(f[0] == '.'):
				pass
			else:
				# ��ӷ������ļ���
				dirList.append(f)
		if(os.path.isfile(path + '/' + f)):
			# ����ļ�
			fileList.append(f)

# �����洢���ݵ��ֵ� 
data_dic = [] 

def updateFile(path, filename):
	'''
	wkfile = path +"/"+ filename

	wb = Workbook()
	#ws = wb.active
	ws = wb.get_sheet_by_name(name='statistics')
	ws.auto_filter.ref = "A1:D4"
	#ws.title = 'statistics'
	
	wb.save(wkfile)
	'''

	wkfile = path +"/"+ filename 
	print wkfile

	wb = load_workbook(wkfile)
	#print "Worksheet range(s):", wb.get_named_ranges()
	#print "Worksheet name(s):", wb.get_sheet_names()
	if sheet_st not in wb.get_sheet_names():
		print "No Sheet"
		return

	ws = wb.get_sheet_by_name(sheet_st)
	ws.auto_filter.ref = "A1:F1"
	wb.save(path +"/bk_"+ filename)

if __name__ == '__main__':
	pindex = len(sys.argv)
	if pindex<2:
		sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " ����\n")
		exit(1);

	code = sys.argv[1]
	if (len(code) != 6):
		sys.stderr.write("Len should be 6\n")
		exit(1);

	head3 = code[0:3]
	result = (cmp(head3, "000")==0) or (cmp(head3, "002")==0) or (cmp(head3, "300")==0)
	if result is True:
		code = "sz" + code
	else:
		result = (cmp(head3, "600")==0) or (cmp(head3, "601")==0) or (cmp(head3, "603")==0)
		if result is True:
			code = "sh" + code
		else:
			print "�Ƿ�����:" +code+ "\n"
			exit(1);

	fileList = []
	path = prepath + code
	if not os.path.isdir(path):
		print "'"+ path +"' not a dir"
		exit(1)

	for (dirpath, dirnames, filenames) in os.walk(path):  
		#print('dirpath = ' + dirpath)
		i = 0
		for filename in filenames:
			extname = filename.split('.')[-1]
			print filename, extname
			if cmp(extname,"xlsx")!=0:
				continue
			prename = code+"__"
			if cmp(filename[0:9], prename[0:9])!=0:
				continue

			print filename
			updateFile(path, filename)
			break

		#�����õ����ļ��е��ļ����������ļ������ļ�
		break;
