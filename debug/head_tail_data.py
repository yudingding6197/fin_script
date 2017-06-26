#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib
import urllib2
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
import tushare as ts
sys.path.append(".")
sys.path.append("..")
from internal.ts_common import *

#某一只每天的开收集合价信息

# Main
dpath = "../Data/"
statis = "../Data/entry/stat/"
pindex = len(sys.argv)
if pindex<2:
	sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " 代码\n")
	exit(1);

code = sys.argv[1]
if (len(code) != 6):
	sys.stderr.write("Len should be 6\n")
	exit(1);

head3 = code[0:3]
result = (cmp(head3, "000")==0) or (cmp(head3, "002")==0) or (cmp(head3, "300")==0)
if result is True:
	code = "sz" + code
else:
	result = (cmp(head3, "600")==0) or (cmp(head3, "601")==0) or (cmp(head3, "603")==0)
	if result is True:
		code = "sh" + code
	else:
		print "非法代码:" +code+ "\n"
		exit(1);
prepath = dpath + code + "/"

stwb = Workbook()
stws = stwb.active

items = os.listdir(prepath)
items.reverse()

strline = u'日期,开价,涨跌幅,成交量,收价,涨跌幅,成交量,收盘价,涨跌幅,前收价,开盘价,最高价,最低价'
strObj = strline.split(u',')
stws.append(strObj)

row = 2
for f in items:
	if(os.path.isdir(prepath + f)):
		continue
	date = f[9:19]
	list = []
	list.append(date)
	#print f,date
	wb = load_workbook(prepath + f)
	sheetnames = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheetnames[0])
	#ws = wb.active

	#print "Work Sheet Titile:", ws.title
	#print "Work Sheet Rows:", ws.max_row
	#print "Work Sheet Cols:", ws.max_column
	if ws.max_column!=15:
		print "Warning: Check the column:", f, ws.max_column
		continue
	c_time = ws.cell(row=2, column=1).value
	c_price = ws.cell(row=2, column=2).value
	c_change = ws.cell(row=2, column=3).value
	c_volumn = ws.cell(row=2, column=5).value
	#print c_time,c_price,c_change,c_volumn

	o_time = ws.cell(row=ws.max_row, column=1).value
	o_price = ws.cell(row=ws.max_row, column=2).value
	o_change = ws.cell(row=ws.max_row, column=3).value
	o_volumn = ws.cell(row=ws.max_row, column=5).value
	#print o_time,o_price,o_change,o_volumn

	fin_c_prc = ws.cell(row=2, column=8).value
	fin_change = ws.cell(row=2, column=9).value
	fin_p_prc = ws.cell(row=2, column=10).value
	fin_o_prc = ws.cell(row=2, column=11).value
	fin_high = ws.cell(row=2, column=12).value
	fin_low = ws.cell(row=2, column=13).value

	list.append(o_price)
	list.append(o_change)
	list.append(o_volumn)

	list.append(c_price)
	list.append(c_change)
	list.append(c_volumn)

	list.append(fin_c_prc)
	list.append(fin_change)
	list.append(fin_p_prc)
	list.append(fin_o_prc)
	list.append(fin_high)
	list.append(fin_low)
	
	if (c_price!=fin_c_prc) or (o_price!=fin_o_prc):
		value = c_price,fin_c_prc,o_price,fin_o_prc
		print "Waring: " + f + " Not equal item:", value

	k = 0
	ascid = 65
	for k in range(0, len(list)):
		cell = chr(ascid+k) + str(row)
		stws[cell] = list[k]
	row += 1

stws.auto_filter.ref = "A1:M1"
filexlsx = statis +code+ '.xlsx'
stwb.save(filexlsx)
filexlsx = dpath + 'hd_tl_trade.xlsx'
stwb.save(filexlsx)
