# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib
import urllib2
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
import ctypes

#url = "http://vip.stock.finance.sina.com.cn/quotes_service/view/vMS_tradedetail.php?symbol=sz300001&date=2015-09-10&page=48"
#成交时间	成交价	涨跌幅	价格变动	成交量(手)		成交额(元)	性质
#	<tr ><th>11:29:48</th><td>14.57</td><td>-3.06%</td><td>+0.01</td><td>16</td><td>23,312</td><th><h1>中性盘</h1></th></tr>
#<tr ><th>11:29:36</th><td>14.56</td><td>-3.13%</td><td>--</td><td>9</td><td>13,104</td><th><h6>卖盘</h6></th></tr>
#<tr ><th>11:29:21</th><td>14.56</td><td>-3.13%</td><td>-0.02</td><td>10</td><td>14,560</td><th><h6>卖盘</h6></th></tr>

class fitItem:
	volumn = 0
	buyvol = 0
	buyct = 0
	buyavg = 0
	sellvol = 0
	sellct = 0
	sellavg = 0
	def __init__(self, vol):
		self.volumn = vol
		self.buyvol = 0
		self.buyct = 0
		self.buyavg = 0
		self.sellvol = 0
		self.sellct = 0
		self.sellavg = 0

dftsarr = "0,100,200,300,600,900"
Handle_Mid = 0
Large_Volume = 2000
Tras_Count = 5
Large_Vol_Time = []

# _____ debug print log
def loginfo(flag=0):
	if (flag==1):
		frame = None
		try:
			raise  ZeroDivisionError
		except  ZeroDivisionError:
			frame = sys.exc_info()[2].tb_frame.f_back
		print "%s in line %d" %(str(datetime.datetime.now()), frame.f_lineno)

#未来对部分全局参数进行设置
def initParam():
	pass

def parseCode(code):
	if (len(code) != 6):
		sys.stderr.write("Len should be 6\n")
		return (-1, '')

	shcd = ['600', '601', '603']
	szcd = ['000','001','002','300']
	head3 = code[0:3]
	if head3 in szcd:
		ncode = "sz" + code
	else:
		if head3 in shcd:
			ncode = "sh" + code
		else:
			print "非法代码:" +code+ "\n"
			return (-1, '')
	return (0, ncode)

def parseDate(qdate, today):
	dateObj = re.match(r'^(\d{4})(\d{2})(\d{2})', qdate)
	if (dateObj is None):
		dateObj = re.match(r'^(\d{2})(\d{2})', qdate)
		if (dateObj is None):
			print "非法日期格式：" +qdate+ ",期望格式:YYYYMMDD or MMDD"
			return (-1, '')
		else:
			if len(qdate)>4:
				print "非法日期格式：" +qdate+ ",期望格式:YYYYMMDD or MMDD"
				return (-1, '')
			year = today.year
			month = int(dateObj.group(1))
			day = int(dateObj.group(2))
	else:
		year = int(dateObj.group(1))
		month = int(dateObj.group(2))
		day = int(dateObj.group(3))
	strdate = '%04d-%02d-%02d' %(year, month, day)

	try:
		datetime.date(year,month,day)
	except:
		print strdate, "is invalid date"
		return (-1, '')
	return (0, strdate)

def parseSeparateDate(qdate, today):
	dateObj = re.match(r'^(\d{4})(\d{2})(\d{2})', qdate)
	if (dateObj is None):
		dateObj = re.match(r'^(\d{2})(\d{2})', qdate)
		if (dateObj is None):
			print "非法日期格式：" +qdate+ ",期望格式:YYYYMMDD or MMDD"
			return (-1, '')
		else:
			year = today.year
			month = int(dateObj.group(1))
			day = int(dateObj.group(2))
	else:
		year = int(dateObj.group(1))
		month = int(dateObj.group(2))
		day = int(dateObj.group(3))

	#验证日期的合法性
	try:
		datetime.date(year,month,day)
	except:
		print year, month, day, "is invalid date"
		return (-1, -1,-1,-1)
	return (0, year,month,day)

def parseTime(qtime):
	timeObj = re.match(r'^(\d{2}):(\d{2}):(\d{2})', qtime)
	if (timeObj is None):
		print "非法时间格式：" +qtime+ ",期望格式:HH:MM:SS"
		return (-1, -1, -1, -1)

	hour = int(timeObj.group(1))
	minute = int(timeObj.group(2))
	second = int(timeObj.group(3))
	return (0, hour, minute, second)

def time_range(firstHour, firstMinute, hour, minute, interval):
	bMatch = 0
	if firstMinute<interval:
		if (hour==firstHour and minute<=firstMinute):
			bMatch = 1
		elif ((firstHour-hour==1) and minute>=(60-interval+firstMinute)):
			bMatch = 1
	else:
		if (hour==firstHour and firstMinute-minute<=5):
			bMatch = 1
	return bMatch

#type:
#	1:Buy,	2:Sell
#当第4个参数设置为1，进行折半处理
#某个时刻的成交量，满足条件时，求和计算总成交量
def handle_volumn(exVolumn, dataObj, type, flag=0):
	dataObjLen = len(dataObj)
	chgVolumn = exVolumn
	if flag==1:
		chgVolumn = exVolumn/2

	if type==1:
		for j in range(0, dataObjLen):
			if exVolumn<dataObj[j].volumn:
				break;
			dataObj[j].buyvol += chgVolumn
			dataObj[j].buyct += 1
	elif type==2:
		for j in range(0, dataObjLen):
			if exVolumn<dataObj[j].volumn:
				break;
			dataObj[j].sellvol += chgVolumn
			dataObj[j].sellct += 1

def handle_middle_volumn(exVolumn, dataObj, exTime, fluctuate, increaseRange):
	if Handle_Mid!=1:
		return 0

	ret = 0
	dataObjLen = len(dataObj)

	#特殊处理第一笔 (9:25) 的成交
	timeobj = exTime.split(':')
	hour = int(timeobj[0])
	minute = int(timeobj[1])
	if (hour==9 and (minute>24 and minute<30)):
		if increaseRange is None:
			print "______ It is NONE"
			return 0
		rangeobj = re.match(r'(.*)\%', increaseRange)
		if rangeobj is None:
			print "Not match string:", increaseRange
			return 0

		fltval = float(rangeobj.group(1))
		if fltval>0.001:
			handle_volumn(exVolumn, dataObj, 1)
		elif fltval<-0.001:
			handle_volumn(exVolumn, dataObj, 2)
		else:
			handle_volumn(exVolumn, dataObj, 1, 1)
			handle_volumn(exVolumn, dataObj, 2, 1)
		return 0

	if (fluctuate=='--'):
		handle_volumn(exVolumn, dataObj, 1, 1)
		handle_volumn(exVolumn, dataObj, 2, 1)
		return 0

	ftfluct = float(fluctuate)
	if (ftfluct>-0.021 and ftfluct<0.021):
		handle_volumn(exVolumn, dataObj, 1, 1)
		handle_volumn(exVolumn, dataObj, 2, 1)
	elif (ftfluct>1.0 or ftfluct<-1.0):
		print "??? Flucture is too big:", fluctuate
	elif (ftfluct>0.02):
		handle_volumn(exVolumn, dataObj, 1)
		ret = 1
	elif (ftfluct<-0.02):
		handle_volumn(exVolumn, dataObj, 2)
		ret = 2
	return ret

def write_statics(ws, fctime, dataObj, qdate, savedTrasData, largeTrasData):
	ws.title = 'statistics'

	ascid = 65
	row = 1
	if cmp(fctime, '')==0:
		title = [qdate, 'B', 'P1', 'S', 'P2', 'B_vol', 'S_vol', 'B_avg', 'S_avg', ]
	else:
		title = [qdate, 'B', 'P1', 'S', 'P2', 'B_vol', 'S_vol', 'B_avg', 'S_avg', fctime]
	number = len(title)
	for i in range(0,number):
		cell = chr(ascid+i) + str(row)
		ws[cell] = title[i]

	dataObjLen = len(dataObj)
	totalBVol = 0
	totalSVol = 0
	for j in range(0, dataObjLen):
		list = []
		list.append(dataObj[j].volumn)
		
		buyvol = dataObj[j].buyvol
		buyct = dataObj[j].buyct
		
		sellvol = dataObj[j].sellvol
		sellct = dataObj[j].sellct

		buyPerc = ''
		sellPerc = ''
		if dataObj[j].volumn==0:
			totalBVol = buyvol
			totalSVol = sellvol
		if totalBVol!=0:
			if dataObj[j].volumn==0:
				buyPerc = round(float(buyvol) * 100 / (totalBVol+totalSVol), 2);
			else:
				buyPerc = round(float(buyvol) * 100 / totalBVol, 2)
		else:
			buyPerc = 0
		if totalSVol!=0:
			if dataObj[j].volumn==0:
				sellPerc = round(float(sellvol) * 100 / (totalBVol+totalSVol), 2)
			else:
				sellPerc = round(float(sellvol) * 100 / totalSVol, 2)
		else:
			sellPerc = 0

		list.append(buyvol)
		list.append(buyPerc)
		list.append(sellvol)
		list.append(sellPerc)
		list.append(buyct)
		list.append(sellct)
		if buyct==0:
			list.append(0)
		else:
			list.append(buyvol/buyct)
		if sellct==0:
			list.append(0)
		else:
			list.append(sellvol/sellct)
		list.append('')
		list.append(buyvol + sellvol)

		percBS=''
		if (totalBVol + totalSVol)!=0:
			percBS = round(float(buyvol + sellvol) * 100 / (totalBVol + totalSVol), 2)
		list.append(percBS)

		list.append(buyct + sellct)
		bsct = buyct + sellct
		if bsct==0:
			list.append(0)
		else:
			list.append((buyvol + sellvol)/bsct)
		
		row = row+1
		number = len(list)
		for i in range(0,number):
			cell = chr(ascid+i) + str(row)
			ws[cell] = list[i]

	#再添加交易数据
	row = row+1
	dataObjLen = len(savedTrasData)
	for j in range(0, dataObjLen):
		row = row+1
		list = savedTrasData[j]
		trasLen = len(list)
		for i in range(0,trasLen):
			cell = chr(ascid+i) + str(row)
			ws[cell] = list[i]

	#再添加大单数据
	dataObjLen = len(largeTrasData)
	if dataObjLen>0:
		row = row+2
		cell = chr(ascid) + str(row)
		ws[cell] = u'大单记录'
		for j in range(0, dataObjLen):
			row = row+1
			list = largeTrasData[j]
			trasLen = len(list)
			for i in range(0,trasLen):
				cell = chr(ascid+i) + str(row)
				ws[cell] = list[i]


def	handle_last_price(tmpContPrice, contPrice):
	tmpPriceLen = len(tmpContPrice)
	contPriceLen = len(contPrice)
	bChange = 0
	#逻辑处理，两个数组，新的数组没有数据就忽略
	#tmpContPrice=['买盘', 1, 23, 789, ...]
	#临时数组前两个值含义: tmpContPrice[0] = '买盘/卖盘', tmpContPrice[1] = '0/1'  允许添加，还是不能再添加了
	#大单顺序 '买','买','买','卖','卖','买','买'
	#'卖'后面的'买单'不能再添加到数组，一共3笔连续大买单
	if tmpPriceLen>2:
		#原来数组没有数据，直接替换
		if contPriceLen==0:
			for k in range(2,tmpPriceLen):
				contPrice.append(tmpContPrice[k])
		else:
			#对比原来数组和临时数组，临时数组变小，则意味数据一定改变了
			if (tmpPriceLen-2)<contPriceLen:
				#tmpLst=['B', 1, 23, 789]
				#conLst=[12, 122, 32]
				bAllMatch = 1
				for k in range(2,tmpPriceLen):
					if (tmpContPrice[k]!=contPrice[k-2]):
						bAllMatch = 0
				if bAllMatch==0:
					contPrice.append(tmpContPrice[k])
					del contPrice[:]
					for k in range(2,tmpPriceLen):
						contPrice.append(tmpContPrice[k])
					bChange = 1
			else:
				#原来数组小于或者等于临时数组，对比每一个数据
				bAllMatch = 1
				for k in range(0, contPriceLen):
					if (tmpContPrice[k+2]!=contPrice[k]):
						bAllMatch = 0
				if bAllMatch==0:
					#如果数据不完全一致，重新替换
					#tmpLst=['B', 1, 12, 111, 55, 343, 455]
					#conLst=[12, 122, 32]
					del contPrice[:]
					for k in range(2,tmpPriceLen):
						contPrice.append(tmpContPrice[k])
					bChange = 1
				else:
					#数据完全一致，意味着在原来数组基础上增加新值
					#tmpLst=['B', 1, 12, 122, 32, 343, 455]
					#conLst=[12, 122, 32]
					if (tmpPriceLen-2)>contPriceLen:
						for k in range(contPriceLen+2,tmpPriceLen):
							contPrice.append(tmpContPrice[k])
						bChange = 1
		#如果数据发生了改变，增加或者重新替换，但是对长度有要求：超过5
		contPriceLen = len(contPrice)
		if (contPriceLen>=6 and bChange==1):
			#print "Price:::",contPrice
			#msgstr = u'Continued value:%d'%(contPriceLen)
			#ctypes.windll.user32.MessageBoxW(0, msgstr, '', 0)
			return 1
	return 0

def handle_data(addcsv, prepath, bhist, url, code, qdate, sarr):
	todayUrl = "http://hq.sinajs.cn/list=" + code
	if bhist==0:
		url = url +"?symbol="+ code
	elif bhist==1:
		url = url +"?date="+ qdate +"&symbol="+ code
	#当天日期按照历史记录处理
	elif bhist==2:
		url = url +"?symbol="+ code
	else:
		print "Unknown flag:", bhist
		return -1
	#if Handle_Mid==0:
	#	print "Message: Ignore 中性盘"

	if not os.path.isdir(prepath):
		os.makedirs(prepath)

	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	dataObjLen = len(dataObj)
	check_vol1 = int(volObj[1])
	check_vol2 = int(volObj[3])
	Large_Volume = int(volObj[arrlen-1])*2

	wb = Workbook()
	# grab the active worksheet
	ws = wb.active

	totalline = 0
	#可能数据在不同的页面，同时存在，这是重复数据需要过滤重复结果
	#还可能相同时间，产生多个成交量，需要都保留
	lasttime = ''
	lastvol = 0
	pageFtime = ''
	bFtime = 0
	bFindHist = 0
	hisUrl = ''
	filename = code+ '_' + qdate
	fctime = ''
	todayData = []
	todayDataLen = 0
	bGetToday = 0

	cur=datetime.datetime.now()
	if bhist==0:
		fctime = '%02d:%02d' %(cur.hour, cur.minute)
		filename = '%s_%02d-%02d' %(filename, cur.hour, cur.minute)
		bGetToday = 1
	if (bhist==2 and cur.hour>=15 and cur.minute>0):
		bGetToday = 1
	if (bGetToday==1):
		try:
			req = urllib2.Request(todayUrl)
			stockData = urllib2.urlopen(req, timeout=5).read()
		except:
			loginfo(1)
			print "URL timeout"
		else:
			stockObj = stockData.split(',')
			if len(stockObj)<10:
				print code, ": No trade data"
				return -1

			closePrice = float(stockObj[3])
			lastClsPrice = float(stockObj[2])
			openPrice = float(stockObj[1])
			highPrice = float(stockObj[4])
			lowPrice = float(stockObj[5])
			exVolume = int(stockObj[8])/100
			exAmount = float(stockObj[9])
			f1 = '%02.02f'%( ((closePrice-lastClsPrice)/lastClsPrice)*100 )
			exFluc = float(f1)
			
			todayData.append(closePrice)
			todayData.append(exFluc)
			todayData.append(lastClsPrice)
			todayData.append(openPrice)
			todayData.append(highPrice)
			todayData.append(lowPrice)
			todayData.append(exVolume)
			todayData.append(exAmount)
			todayDataLen = len(todayData)
			#print todayData

	if addcsv==1:
		filecsv = prepath + filename + '.csv'
		fcsv = open(filecsv, 'w')
		strline = '成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质'
		fcsv.write(strline)
		fcsv.write("\n")

	strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质,收盘价,涨跌幅,前收价,开盘价,最高价,最低价,成交量,成交额'
	strObj = strline.split(u',')
	ws.append(strObj)
	#dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D')
	dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+|\+?\d+|\-\d+)\D+(\d+)</td><td>([\d,]+)</td><th>(.*)\D')
	frameRe = re.compile(r'.*name=\"list_frame\" src=\"(.*)\" frameborder')
	keyw = '收盘价|涨跌幅|前收价|开盘价|最高价|最低价|成交量|成交额'
	infoRe = re.compile(r'\D+('+keyw+').*>(\+?-?\d+\.\d+)')
	excecount = 0
	stockInfo = []
	reloadUrl = 0
	noDataFlag = 0
	noDataKey = "该股票没有交易数据"
	notTrasFlag = 0
	notTrasKey = "输入的日期为非交易日期"
	#每一页的数据，如果找到匹配数据则设置为1；解决有时候页面有数据但是收不到，
	#count为0，重新加载尝试再次获取；如果解析到数据的页面，如果count为0就不再继续解析数据
	matchDataFlag = 0
	savedTrasData = []
	savedTrasData2 = []
	largeTrasData = []
	i = 1
	lineCount = 0
	firstRec = 0

	for j in range(1,1000):
		urlall = url + "&page=" +str(i)
		#print "(%d):%s" %(i,urlall)

		if excecount>10:
			print "Quit with exception i=", i
			break

		#创建url链接，获取每一页的数据
		try:
			req = urllib2.Request(urlall)
			res_data = urllib2.urlopen(req, timeout=5).readlines()
			lineCount = len(res_data)
		except:
			print "URL1 except:",urlall
			excecount += 1
			continue
		else:
			excecount = 0
			pass
		if lineCount==0:
			break

		#print "(%d):%s FIN" %(i,url)
		flag = 0
		count = 0
		bFtime = 0
		matchDataFlag = 0
		idx = 0
		line = ''
		firstRec = 0
		
		#开始读取每一页返回的内容，首先查找'成交时间'/'收盘价'，过滤大量不需要的内容
		if bhist==0 or bhist==2:
			checkStr = '成交时间'
		else:
			checkStr = '收盘价'

		#查找到'成交时间'/'收盘价'，更新查找内容为'<script type='
		#但是一个月前的历史记录，'<script type='不是找到，需要handle_his_data()处理
		while True:
			if idx>=lineCount:
				break
			line = res_data[idx]
			idx += 1
			#print line
			index = line.find(checkStr)
			if (index>=0):
				checkStr = '<script type='
				break;
		if idx>=lineCount:
			break

		#从第一页读取：收盘价,涨跌幅,前收价等数据
		if bhist==1 and i==1:
			infoCount = 0
			while True:
				infoObj = infoRe.match(line)
				if infoObj:
					stockInfo.append(float(infoObj.group(2)))
				else:
					if infoCount<8:
						print "Parse fail", line
					break;
				infoCount += 1
				line = res_data[idx]
				idx += 1
			#print stockInfo

		while True:
			line = res_data[idx]
			idx += 1
			#print line
			index = line.find(checkStr)
			if (index>=0):
				#找到关键字'<script type='，找到则后面的数据不用分析，准备获取下一页的内容了
				#print "%s found, QUIT line='%s' "%(checkStr, line)
				break;

			#key = re.match(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D', line)
			key = dtlRe.match(line)
			if key:
				#print key.groups(), sys._getframe().f_lineno 
				curtime = key.group(1)
				curvol = int(key.group(5))
				if curvol==0:
					continue

				ret,hour,minute,second = parseTime(curtime)
				if (ret==-1):
					continue

				#记住当前页第一个的时间
				if (bFtime==0):
					timeobj = re.search(curtime, pageFtime)
					if timeobj:
						print "bFtime matched ", curtime
						if (hour==9 and minute<30):
							firstRec = 1
						break
					pageFtime = curtime
					bFtime = 1

				matchDataFlag = 1
				if (key.group(2)=="0.00") or (key.group(3)=="-100.00%"):
					print "page(%d) Price(%s) or range(%s) is invalid value"%(i, key.group(2), key.group(3))
					continue
				if (hour==9 and minute<=20) or (hour==15 and minute>1):
					count += 1
					continue

				timeobj = re.search(curtime, lasttime)
				if (timeobj and curvol==lastvol):
					count += 1
					if (hour==9 and minute>20 and minute<=26 and count==1):
						noDataFlag = 1
					continue

				#此时这个if判断没有意义了，前面代码做了判断
				if (key.group(2)=="0.00") or (key.group(3)=="-100.00%"):
					print "page(%d) Price(%s) or range(%s) is invalid value"%(i, key.group(2), key.group(3))
				else:
					volume = int(key.group(5))
					curprice = key.group(2)
					p_change = key.group(3)
					obj = re.match (r'([+-]?\d{1,3}.\d{1,3})\%', p_change)
					if obj is None:
						print "Warning: Invalid change ", curtime, p_change
						continue
					change = float(obj.group(1))
					fluctuate = key.group(4)
					lasttime = curtime
					lastvol = curvol
					amount = key.group(6)
					obj = amount.split(',')
					amount = ''.join(obj)

					state = key.group(7)
					if state[0:2]=='--':
						state = '中性盘'
					else:
						bsArray = re.match(r'<h\d+>(卖盘|买盘|中性盘)\D', state)
						state = bsArray.group(1)

					stateStr = ''
					bAddVolumn = 1
					if (hour==9 and minute==25) or (hour==15 and minute==0):
						bAddVolumn = 0

					stateStr = state
					if cmp(state, '卖盘')==0:
						stateStr = 'SELL卖盘'
						if bAddVolumn==1:
							handle_volumn(volume, dataObj, 2)
					elif cmp(state, '买盘')==0:
						if bAddVolumn==1:
							handle_volumn(volume, dataObj, 1)
					#目前中性盘没有处理
					elif cmp(state, '中性盘')==0:
						if bAddVolumn==1:
							ret = handle_middle_volumn(volume, dataObj, curtime, fluctuate, change)
						else:
							ret = 0
						if ret==1:
							stateStr = '买盘'
						elif ret==2:
							stateStr = 'SELL卖盘'

					if addcsv==1:
						strline = curtime +","+ key.group(2) +","+ key.group(3) +","+ key.group(4) +","+ key.group(5) +","+ amount +","+ stateStr + "\n"
						fcsv.write(strline)

					totalline += 1
					row = totalline+1
					price = float(key.group(2))
					cell = 'A' + str(row)
					ws[cell] = curtime
					cell = 'B' + str(row)
					ws[cell] = price
					cell = 'C' + str(row)
					ws[cell] = change
					cell = 'D' + str(row)
					ftfluct = fluctuate
					if (fluctuate=='--'):
						ws[cell] = key.group(4)
					else:
						ftfluct = float(fluctuate)
						ws[cell] = ftfluct
					cell = 'E' + str(row)
					ws[cell] = curvol
					cell = 'F' + str(row)
					ws[cell] = int(amount)
					cell = 'G' + str(row)
					s1 = stateStr.decode('gbk')
					ws[cell] = s1

					#将当天的数据在Sheet页面更新
					if (row==2 and (bhist==0 or (bhist==2 and cur.hour>=15)) and todayDataLen>0):
						ascid = 72
						for k in range(0, todayDataLen):
							cell = chr(ascid+k) + str(row)
							ws[cell] = todayData[k]

					#将开始和最后成交数据保存
					bSaveFlag = 0
					if (totalline==1 or (totalline<4 and curvol>check_vol1)):
						bSaveFlag = 1
					elif (hour==9 and minute==30 and curvol>check_vol2) or (hour==9 and minute<30):
						bSaveFlag = 2
					if bSaveFlag==1 or bSaveFlag==2:
						rowData = []
						rowData.append(curtime)
						rowData.append(price)
						rowData.append(change)
						rowData.append(ftfluct)
						rowData.append(curvol)
						rowData.append(int(amount))
						rowData.append(s1)
					if bSaveFlag==1:
						savedTrasData.append(rowData)
					elif bSaveFlag==2:
						savedTrasData2.append(rowData)

					#增加大单成交记录
					if (curvol>=Large_Volume):
						rowData = []
						rowData.append(curtime)
						rowData.append(price)
						rowData.append(change)
						rowData.append(ftfluct)
						rowData.append(curvol)
						rowData.append(int(amount))
						rowData.append(s1)
						largeTrasData.append(rowData)

					if (row==2 and bhist==1):
						ascid = 72
						number = len(stockInfo)
						for k in range(0,number):
							cell = chr(ascid+k) + str(row)
							ws[cell] = stockInfo[k]

				count += 1
				continue
			else:
				index = line.find(noDataKey)
				if (index>=0):
					#找到关键字，当前和以后的页面都没有数据
					#print "KEY word '%s' found, QUIT line='%s' "%(noDataKey, line)
					noDataFlag = 1
					break;

				index = line.find(notTrasKey)
				if (index>=0):
					#找到关键字，非交易日
					#print "KEY word '%s' found, QUIT line='%s' "%(notTrasKey, line)
					notTrasFlag = 1
					break;

				#针对一个月前的历史记录，找到匹配的关键字
				if bhist==1:
					key = frameRe.match(line)
					if key:
						bFindHist = 1
						hisUrl = key.group(1)
						#print key.groups()
						break

		if bFindHist==1:
			break

		#通过此方法判断是否还有数据
		if (noDataFlag==1):
			#print "No data found current page=", i, ", QUIT"
			break

		#通过此方法判断是否非交易日
		if (notTrasFlag==1):
			#print "No Tras found current page=", i, ", QUIT"
			break

		#此时还应该有数据，但是得到的数据数量为0，重新获取数据
		if (count==0):
			if firstRec==1:
				break
			if totalline==1:
				print "Warning: Only one line data"
				print urlall
				break
			if (matchDataFlag==1):
				print "Warnig: All invalid data in page=", i
				break
			print "Warnig: !!! Reload data in page=", i
			reloadUrl += 1
			if (reloadUrl>9):
				print "Get Data:获取数据可能不完整，建议重新获取"
				break
			continue

		#最后i加一，访问下一页，对应 for 循环启动代码
		i += 1
	ws.auto_filter.ref = "A1:G1"
	#ws.auto_filter.add_filter_column(4, ['300'])
	#ws.auto_filter.filter_columns()

	if addcsv==1:
		fcsv.close()
		if (totalline==0):
			os.remove(filecsv)

	if totalline>0:
		startIdx = 0
		savedTrasLen = len(savedTrasData2)
		if savedTrasLen>0:
			if savedTrasLen>Tras_Count:
				startIdx = savedTrasLen-Tras_Count
			for j in range(startIdx, savedTrasLen):
				savedTrasData.append(savedTrasData2[j])
		ws = wb.create_sheet()
		write_statics(ws, fctime, dataObj, qdate, savedTrasData, largeTrasData)

	filexlsx = prepath +filename+ '.xlsx'
	if (os.path.exists(filexlsx) and bhist==0):
		j = 1
		while True:
			filexlsx = prepath + filename + '_' + str(j) + '.xlsx'
			j += 1
			if not os.path.exists(filexlsx):
				break;

	if (totalline==0):
		filexlsx = prepath +filename+ '_tmp0.xlsx'
		wb.save(filexlsx)
		os.remove(filexlsx)
		if bFindHist==1:
			handle_his_data(addcsv, prepath, hisUrl, code, qdate, stockInfo, sarr)
		else:
			print qdate+ " Handle data, No Matched Record"
	else:
		wb.save(filexlsx)
		#loginfo()
		return 0
	return -1

#处理要求跳转页面的历史记录
def handle_his_data(addcsv, prepath, url, code, qdate, stockInfo, sarr):
	#print "handle_his_data, url=",url
	if not os.path.isdir(prepath):
		os.makedirs(prepath)

	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	dataObjLen = len(dataObj)
	check_vol1 = int(volObj[1])
	check_vol2 = int(volObj[3])
	Large_Volume = int(volObj[arrlen-1])*2

	wb = Workbook()
	# grab the active worksheet
	ws = wb.active

	totalline = 0
	#可能数据在不同的页面，同时存在，这是重复数据需要过滤重复结果
	#还可能相同时间，产生多个成交量，需要都保留
	lasttime = ''
	lastvol = 0
	pageFtime = ''
	bFtime = 0
	filename = code+ '_' + qdate
	if addcsv==1:
		filecsv = prepath + filename + '.csv'
		fcsv = open(filecsv, 'w')
		strline = '成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质'
		fcsv.write(strline)
		fcsv.write("\n")

	strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质,收盘价,涨跌幅,前收价,开盘价,最高价,最低价,成交量,成交额'
	strObj = strline.split(u',')
	ws.append(strObj)
	dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(--|\+?\d+.\d+|-\d+.\d+|\+?\d+|\-\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D')
	excecount = 0
	savedTrasData = []
	savedTrasData2 = []
	largeTrasData = []
	lineCount = 0
	i = 1
	for j in range(1,1000):
		urlall = url + "&page=" +str(i)
		#print "HIS(%d):%s" %(i,urlall)

		if excecount>10:
			print "Quit with exception i=", i
			break

		#创建url链接，获取每一页的数据
		try:
			req = urllib2.Request(urlall)
			res_data = urllib2.urlopen(req, timeout=5).readlines()
			lineCount = len(res_data)
		except:
			print "URL2 except=",urlall
			excecount += 1
			continue
		else:
			excecount = 0
			pass
		#print "HIS(%d):%s FIN" %(i,url)

		flag = 0
		count = 0
		bFtime = 0
		idx = 0
		line = ''
		checkStr = '成交时间'
		while True:
			#print line
			if idx>=lineCount:
				break
			line = res_data[idx]
			idx += 1
			index = line.find(checkStr)
			if (index<0):
				continue

			if flag==0:
				checkStr = '<th>'
				flag = 1
			else:
				key = dtlRe.match(line)
				if (key):
					curtime = key.group(1)
					price = key.group(2)
					fluctuate = key.group(3)
					curvol = int(key.group(4))
					intcurvol = int(key.group(4))
					amount = key.group(5)
					state = key.group(6)
					srange = ''

					if intcurvol==0:
						continue

					#记住当前页第一个的时间
					if (bFtime==0):
						timeobj = re.search(curtime, pageFtime)
						if timeobj:
							break
						pageFtime = curtime
						bFtime = 1

					ret,hour,minute,second = parseTime(curtime)
					if (ret==-1):
						continue

					timeobj = re.search(curtime, lasttime)
					if (timeobj and intcurvol==lastvol):
						pass
					else:
						lasttime = curtime
						lastvol = intcurvol
						obj = amount.split(',')
						amount_n = ''.join(obj)
						intamount = int(amount_n)

						stateStr = state
						if cmp(state, '卖盘')==0:
							stateStr = 'SELL卖盘'
							handle_volumn(intcurvol, dataObj, 2)
						elif cmp(state, '买盘')==0:
							handle_volumn(intcurvol, dataObj, 1)
						elif cmp(state, '中性盘')==0:
							ret = handle_middle_volumn(intcurvol, dataObj, curtime, fluctuate, key.group(3))
							if ret==1:
								stateStr = '买盘'
							elif ret==2:
								stateStr = 'SELL卖盘'

						if addcsv==1:
							strline = curtime +","+ price +","+ srange +","+ fluctuate +","+ curvol +","+ amount_n +","+ stateStr +"\n"
							fcsv.write(strline)

						totalline += 1
						row = totalline+1
						cell = 'A' + str(row)
						ws[cell] = curtime
						cell = 'B' + str(row)
						ws[cell] = float(price)
						cell = 'C' + str(row)
						ws[cell] = srange
						cell = 'D' + str(row)
						ws[cell] = fluctuate
						cell = 'E' + str(row)
						ws[cell] = curvol
						cell = 'F' + str(row)
						ws[cell] = intamount
						cell = 'G' + str(row)
						s1 = stateStr.decode('gbk')
						ws[cell] = s1

						if row==2:
							ascid = 72
							number = len(stockInfo)
							for k in range(0,number):
								cell = chr(ascid+k) + str(row)
								ws[cell] = stockInfo[k]

						#将开始和最后成交数据保存
						bSaveFlag = 0
						if (totalline==1 or (totalline<4 and curvol>check_vol1)):
							bSaveFlag = 1
						elif (hour==9 and minute==30 and curvol>check_vol2) or (hour==9 and minute<30):
							bSaveFlag = 2
						if bSaveFlag==1 or bSaveFlag==2:
							rowData = []
							rowData.append(curtime)
							rowData.append(float(price))
							rowData.append(srange)
							rowData.append(fluctuate)
							rowData.append(curvol)
							rowData.append(intamount)
							rowData.append(s1)
						if bSaveFlag==1:
							savedTrasData.append(rowData)
						elif bSaveFlag==2:
							savedTrasData2.append(rowData)

						#增加大单成交记录
						if (curvol>=Large_Volume):
							rowData = []
							rowData.append(curtime)
							rowData.append(price)
							rowData.append(key.group(3))
							rowData.append(fluctuate)
							rowData.append(curvol)
							rowData.append(intamount)
							rowData.append(s1)
							largeTrasData.append(rowData)

					count += 1
				else:
					endObj = re.search(r'</td><td>', qdate)
					if (endObj):
						print "Error line:" + line
					else:
						break;

		#如果没有任何数据得到，表示该页没有数据，以后页也都没有数据，退出循环
		if (count==0):
			print "No data found i=", i, ", QUIT"
			break;
		#最后i加一，访问下一页，对应 for 循环启动代码
		i += 1

	ws.auto_filter.ref = "A1:G1"

	if addcsv==1:
		fcsv.close()
		if (totalline==0):
			os.remove(filecsv)

	if totalline>0:
		startIdx = 0
		savedTrasLen = len(savedTrasData2)
		if savedTrasLen>0:
			if savedTrasLen>Tras_Count:
				startIdx = savedTrasLen-Tras_Count
			for j in range(startIdx, savedTrasLen):
				savedTrasData.append(savedTrasData2[j])
		ws = wb.create_sheet()
		write_statics(ws, '', dataObj, qdate, savedTrasData, largeTrasData)

	if (totalline==0):
		print qdate +" History data, No Matched Record!"
		filexlsx = prepath +filename+ '_tmp.xlsx'
		wb.save(filexlsx)
		os.remove(filexlsx)
	else:
		filexlsx = prepath +filename+ '.xlsx'
		wb.save(filexlsx)
		#print qdate+ " Saved OK!"

def analyze_data(url, code, sarr, priceList, contPrice):
	url = url +"?symbol="+ code
	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	dataObjLen = len(dataObj)
	Large_Volume = int(volObj[arrlen-1])*2

	totalline = 0
	#可能数据在不同的页面，同时存在，这是重复数据需要过滤重复结果
	#还可能相同时间，产生多个成交量，需要都保留
	lasttime = ''
	lastvol = 0
	pageFtime = ''
	bFtime = 0
	hisUrl = ''
	fctime = ''
	todayData = []
	todayDataLen = 0
	bGetToday = 0

	cur=datetime.datetime.now()

	strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质,收盘价,涨跌幅,前收价,开盘价,最高价,最低价,成交量,成交额'
	strObj = strline.split(u',')
	#dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D')
	dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th>(.*)\D')
	frameRe = re.compile(r'.*name=\"list_frame\" src=\"(.*)\" frameborder')
	keyw = '收盘价|涨跌幅|前收价|开盘价|最高价|最低价|成交量|成交额'
	infoRe = re.compile(r'\D+('+keyw+').*>(\+?-?\d+\.\d+)')
	excecount = 0
	stockInfo = []
	reloadUrl = 0
	noDataFlag = 0
	noDataKey = "该股票没有交易数据"
	notTrasFlag = 0
	notTrasKey = "输入的日期为非交易日期"
	#每一页的数据，如果找到匹配数据则设置为1；解决有时候页面有数据但是收不到，
	#count为0，重新加载尝试再次获取；如果解析到数据的页面，如果count为0就不再继续解析数据
	matchDataFlag = 0
	i = 1
	lineCount = 0
	firstHour = 0
	firstMinute = 0
	firstSecond = 0
	minValue = 0
	maxValue = 0
	curValue = 0
	tmpContPrice = []
	pageIdx = 1
	bAlert = 0
	bLastAlert = 0

	#临时处理方案，对国债逆回购
	head5 = code[0:5]
	bBigChange = (cmp(head5, "sz131")==0) or (cmp(head5, "sh204")==0)
	for j in range(pageIdx,1000):
		urlall = url + "&page=" +str(i)

		#print "(%d):%s" %(i,urlall)
		if excecount>10:
			print "Quit with exception i=", i
			break

		#创建url链接，获取每一页的数据
		try:
			req = urllib2.Request(urlall)
			res_data = urllib2.urlopen(req, timeout=2).readlines()
			lineCount = len(res_data)
		except:
			print "URL3 except:",urlall
			excecount += 1
			continue
		else:
			excecount = 0
			pass
		if lineCount==0:
			break
		#print "(%d):%s FIN" %(i,url)

		flag = 0
		count = 0
		bFtime = 0
		matchDataFlag = 0
		idx = 0
		line = ''
		
		#开始读取每一页返回的内容，首先查找'成交时间'/'收盘价'，过滤大量不需要的内容
		checkStr = '成交时间'

		#查找到'成交时间'/'收盘价'，更新查找内容为'<script type='
		while True:
			if idx>=lineCount:
				break
			line = res_data[idx]
			idx += 1
			#print line
			index = line.find(checkStr)
			if (index>=0):
				checkStr = '<script type='
				break;
		if idx>=lineCount:
			break

		while True:
			line = res_data[idx]
			idx += 1
			#print line
			index = line.find(checkStr)
			if (index>=0):
				#找到关键字'<script type='，找到则后面的数据不用分析，准备获取下一页的内容了
				#print "%s found, QUIT line='%s' "%(checkStr, line)
				break;

			#key = re.match(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D', line)
			key = dtlRe.match(line)
			if key:
				#print key.groups(), sys._getframe().f_lineno 
				curtime = key.group(1)
				curvol = int(key.group(5))
				#记住当前页第一个的时间
				if (bFtime==0):
					timeobj = re.search(curtime, pageFtime)
					if timeobj:
						print "bFtime 0 quit"
						break
					pageFtime = curtime
					bFtime = 1
					#第一条记录
					if j==pageIdx:
						ret,firstHour,firstMinute,firstSecond = parseTime(pageFtime)
						if (ret==-1):
							print "Error：Get First Time Fail"
							break
						#print "First=",pageFtime
						curValue = float(key.group(2))
						minValue = int(curValue*100)
						maxValue = minValue

				matchDataFlag = 1
				ret,hour,minute,second = parseTime(curtime)
				if (ret==-1):
					continue
				if (key.group(2)=="0.00") or (key.group(3)=="-100.00%"):
					print "page(%d) Price(%s) or range(%s) is invalid value"%(i, key.group(2), key.group(3))
					continue
				if (hour==9 and minute<=20) or (hour==15 and minute>1):
					count += 1
					continue

				timeobj = re.search(curtime, lasttime)
				if (timeobj and curvol==lastvol):
					count += 1
					if (hour==9 and minute>20 and minute<=26 and count==1):
						noDataFlag = 1
					continue

				#此时这个if判断没有意义了，前面代码做了判断
				if (key.group(2)=="0.00") or (key.group(3)=="-100.00%"):
					print "page(%d) Price(%s) or range(%s) is invalid value"%(i, key.group(2), key.group(3))
				else:
					curprice = key.group(2)
					fluctuate = key.group(4)
					lasttime = curtime
					lastvol = curvol
					amount = key.group(6)
					obj = amount.split(',')
					amount = ''.join(obj)

					volume = int(key.group(5))
					state = key.group(7)
					if state[0:2]=='--':
						state = '中性盘'
					else:
						bsArray = re.match(r'<h\d+>(卖盘|买盘|中性盘)\D', state)
						state = bsArray.group(1)

					if curvol>Large_Volume:
						bFind = 0
						for k in range(0, len(Large_Vol_Time)):
							if (curtime==Large_Vol_Time[k]):
								bFind=1
								break
						if bFind==0:
							Large_Vol_Time.append(curtime)
							sv = ''
							if cmp(state, '卖盘')==0:
								sv = 'S'
							elif cmp(state, '买盘')==0:
								sv = 'B'
							elif cmp(state, '中性盘')==0:
								sv = 'M'
							#15分钟内的大单
							bMatch = time_range(firstHour, firstMinute, hour, minute, 15)
							if bMatch==1:
								if not bBigChange:
									msgstr = u'Hello Big_DT (%s	%s:%d)'%(curtime, sv, curvol)
									ctypes.windll.user32.MessageBoxW(0, msgstr, '', 0)
					if curvol>=check_vol2:
						if (len(tmpContPrice)==0):
							if (state=='卖盘' or state=='买盘'):
								tmpContPrice.append(state)
								tmpContPrice.append(1)
								tmpContPrice.append(curvol)
						else:
							stateValue = tmpContPrice[0]
							allowAdd = tmpContPrice[1]
							if (allowAdd==1):
								if cmp(stateValue, state)==0:
									tmpContPrice.append(curvol)
								else:
									if cmp(state, '卖盘')==0:
										tmpContPrice[1] = 0
									elif cmp(state, '买盘')==0:
										tmpContPrice[1] = 0
									elif cmp(state, '中性盘')==0:
										tmpContPrice.append(curvol)
						bAlert = handle_last_price(tmpContPrice, contPrice)
						if bLastAlert==1 and bAlert==0:
							print "Price=", contPrice
							totalVol = 0
							contPriceLen = len(contPrice)
							for k in range(0, contPriceLen):
								totalVol += contPrice[k]
							sv = 'BB'
							if tmpContPrice[0]=='卖盘':
								sv = 'SS'
							if not bBigChange:
								msgstr = u'Continued(%s) %d: %d'%(sv, contPriceLen, totalVol/contPriceLen)
								ctypes.windll.user32.MessageBoxW(0, msgstr, '', 0)
							bLastAlert = bAlert
						elif bAlert==1 and bLastAlert==0:
							bLastAlert = bAlert

					totalline += 1
					price = float(key.group(2))
					ftfluct = fluctuate
					if (fluctuate=='--'):
						pass
					else:
						ftfluct = float(fluctuate)

					#针对前5分钟的数据，检查得到最大和最小值
					interval = 5
					bMatch = time_range(firstHour, firstMinute, hour, minute, interval)
					if bMatch==1:
						curIntP = int(float(price)*100)
						if curIntP<minValue:
							minValue = curIntP
						elif curIntP>maxValue:
							maxValue = curIntP
					#else:
					#	break

				count += 1
				continue
			else:
				index = line.find(noDataKey)
				if (index>=0):
					#找到关键字，当前和以后的页面都没有数据
					#print "KEY word '%s' found, QUIT line='%s' "%(noDataKey, line)
					noDataFlag = 1
					break;

				index = line.find(notTrasKey)
				if (index>=0):
					#找到关键字，非交易日
					#print "KEY word '%s' found, QUIT line='%s' "%(notTrasKey, line)
					notTrasFlag = 1
					break;

		#通过此方法判断是否还有数据
		if (noDataFlag==1):
			#print "No data found current page=", i, ", QUIT"
			break

		#通过此方法判断是否非交易日
		if (notTrasFlag==1):
			#print "No data found current page=", i, ", QUIT"
			break

		#此时还应该有数据，但是得到的数据数量为0，重新获取数据
		if (count==0):
			if totalline==1:
				print "Warning: Only one line data"
				break
			if (matchDataFlag==1):
				#print "Warnig: All invalid data in page=", i
				break
			print "Warnig: !!! Reload data in page=", i
			reloadUrl += 1
			if (reloadUrl>9):
				print "获取数据可能不完整，建议重新获取"
				break
			continue

		#最后i加一，访问下一页，对应 for 循环启动代码
		i += 1

	#5分钟之内的大小值
	priceList[0] = curValue
	priceList[1] = int(curValue*100)
	priceList[2] = minValue
	priceList[3] = maxValue
	#print priceList
