#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib2
import zlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from xml.etree import ElementTree as ET
from common import *
from decimal import Decimal
import platform
import ctypes
import tushare as ts
import time
from internal.global_var import *

#reload(sys)
#sys.setdefaultencoding('gbk')

STK_ZT = 1<<0
STK_DT = 1<<1
STK_ZTHL = 1<<2
STK_DTFT = 1<<3
STK_YZZT = 1<<4
STK_YZDT = 1<<5
STK_OPEN_ZT = 1<<6
STK_OPEN_DT = 1<<7
STK_ST_YZZT = 1<<8
STK_ST_YZDT = 1<<9
trade_data= None

class statisticsItem:
	s_date = ''
	s_zt = 0
	s_dt = 0
	s_zthl = 0
	s_dtft = 0
	s_yzzt = 0
	s_yzdt = 0
	s_cxzt = 0
	s_cxdt = 0
	s_open_zt = 0		#涨停开盘
	s_close_zt = 0		#涨停开盘个股收盘仍涨停
	s_open_T_zt = 0		#涨停开盘个股收盘仍涨停，非一字
	s_zt_o_gt_c = 0		#触及涨停,开盘价高于收盘价
	s_dk_zt = 0			#开盘涨停，收盘打开涨停
	s_open_dt = 0		#跌停开盘
	s_st_yzzt = 0
	s_st_yzdt = 0
	s_cx_yz = 0
	s_non_cx_yz = 0
	s_open_sz = 0		#开盘上涨
	s_open_xd = 0		#开盘下跌
	s_open_pp = 0		#开盘平盘
	s_open_dz = 0		#开盘大涨
	s_open_dd = 0		#开盘大跌
	s_close_sz = 0		#收盘上涨
	s_close_xd = 0		#收盘下跌
	s_close_pp = 0		#收盘平盘
	s_close_dz = 0		#收盘大涨
	s_close_dd = 0		#收盘大跌
	s_high_zf = 0		#最高涨幅
	s_low_df = 0		#最低跌幅
	s_cx_yzzt = 0		#次新YZZT
	s_open_dt_dk = 0	#开盘跌停打开
	s_new = 0			#上市新股
	s_total = 0			#总计所有交易票
	s_sw_zt = 0			#上午ZT
	s_xw_zt = 0			#下午ZT
	lst_kd = []			#坑爹个股
	lst_nb = []			#NB，振幅大而且涨幅大
	lst_zt_nb = []			#NB，振幅大而且涨幅大，且ZT
	lst_non_zt_nb = []		#NB，振幅大而且涨幅大，非ZT
	lst_jc = []			#韭菜了，严重坑人
	lst_dt_jc = []			#JC，振幅大而且跌幅大，且ZT
	lst_non_dt_jc = []		#JC，振幅大而且跌幅大，非ZT
	lst_non_yzcx_zt = []		#非次新涨停
	lst_non_yzcx_yzzt = []		#非次新一字涨停
	lst_non_yzcx_zthl = []		#非次新涨停回落
	lst_dt = []					#跌停
	lst_yzdt = []				#YZ跌停
	lst_dtft = []				#跌停反弹
	def __init__(self):
		self.s_zt = 0
		self.s_dt = 0
		self.s_zthl = 0
		self.s_dtft = 0
		self.s_yzzt = 0
		self.s_yzdt = 0
		self.s_cxzt = 0
		self.s_cxdt = 0
		self.s_open_zt = 0
		self.s_close_zt = 0
		self.s_open_T_zt = 0
		self.s_zt_o_gt_c = 0
		self.s_dk_zt = 0
		self.s_open_dt = 0
		self.s_st_yzzt = 0
		self.s_st_yzdt = 0
		self.s_open_sz = 0
		self.s_open_xd = 0
		self.s_open_pp = 0
		self.s_open_dz = 0
		self.s_open_dd = 0
		self.s_close_sz = 0
		self.s_close_xd = 0
		self.s_close_pp = 0
		self.s_close_dz = 0
		self.s_close_dd = 0
		self.s_high_zf = 0
		self.s_low_df = 0
		self.s_cx_yzzt = 0
		self.s_open_dt_dk = 0
		self.s_new = 0
		self.s_total = 0
		self.s_sw_zt = 0
		self.s_xw_zt = 0
		self.lst_kd = []
		self.lst_nb = []
		self.lst_jc = []
		self.lst_non_yzcx_zt = []
		self.lst_non_yzcx_yzzt = []
		self.lst_non_yzcx_zthl = []
		self.lst_dt = []
		self.lst_yzdt = []
		self.lst_dtft = []

def spc_round1(value,bit):
	delen = len(str(value).split('.')[1])
	if delen==4:
		b = int(value*10000)%100
	else:
		b = int(value*1000)%10

	rd_val=float( '{:.2f}'.format(Decimal(str(value))) )
	if b==5:
		if int(value*100)%2==0:
			rd_val+=0.01
	return round(rd_val,2)

def spc_round(value,bit):
	b = int(value*10000)
	b1 = b+51
	j = b1/100
	rd_val = float(j)/100
	return rd_val

#从文件中读取某一天是否是交易日
def init_trade_obj():
	global trade_data
	trade_data = pd.read_csv("internal/trade_holiday.csv")
def get_trade_obj():
	global trade_data
	return trade_data

def chk_holiday(date):
	'''
	判断是否为交易日，返回True or False
	'''
	df = get_trade_obj()
	if df is None:
		print "Error: Not get trade Object from file"
		return False
	holiday = df[df.isOpen == 0]['calendarDate'].values
	if isinstance(date, str):
		tsToday = datetime.datetime.strptime(date, '%Y-%m-%d')
	else:
		print "Error: not string type", type(date)
		return False

	#去掉月和日前面的'0'
	obj = re.match(r'^(\d{4})-(\d{2})-(\d{2})', date)
	year = int(obj.group(1))
	month = int(obj.group(2))
	day = int(obj.group(3))
	ndate1 = '%4d-%d-%d' %(year, month, day)
	ndate2 = '%4d/%d/%d' %(year, month, day)

	bhol = ndate1 in holiday or ndate2 in holiday
	if tsToday.isoweekday() in [6, 7] or bhol:
		return True
	else:
		return False

#准备在第一页插入成交数据，在第一次出现vol_array的地方
MAX_COL = 15
vol_array = [30, 60, 100, 200, 300, 500]

def get_range(c_volumn, vol_array):
	length = len(vol_array)
	if c_volumn<vol_array[0]:
		return -1
	if c_volumn>=vol_array[length-1]:
		return length-1
	for i in range(0, length):
		if c_volumn>=vol_array[i] and c_volumn<vol_array[i+1]:
			return i
	return -1

#满足条件的row，添加数据
def update_row_data(ws, f):
	if ws.max_column!=MAX_COL:
		print "Warning: Check the column:", f, ws.max_column
		return
	if ws.max_row < 2:
		print "Warning: max row less than 2>", f, ws.max_column
		return

	marked = [0, 0, 0, 0, 0, 0]
	index = 2
	c_volumn = ws.cell(row=index, column=5).value

	fin_list = []
	for j in range(8, MAX_COL+1):
		item = ws.cell(row=index, column=j).value
		if item is None or item=='':
			fin_list.append(item)
			continue
		fin_list.append(round(item, 2))
	#print fin_list

	#找到符合条件的row，并记录行号
	for index in range(2, ws.max_row):
		c_volumn = ws.cell(row=index, column=5).value
		pos = get_range(c_volumn, vol_array)
		if pos<-1:
			continue
		if marked[pos]==0:
			marked[pos]=index
	#最后一行加上数据，如果最后一行vol太小，还加上>vol_array[0]的行
	mx_row = ws.max_row
	marked.append(mx_row)
	if ws.cell(row=mx_row, column=5).value<vol_array[0]:
		index = mx_row
		while index>=3:
			c_volumn = ws.cell(row=index, column=5).value
			if c_volumn>=vol_array[0]:
				marked.append(index)
				break;
			index -= 1

	for value in marked:
		if value==0:
			continue
		k = 7
		ascid = 65
		for item in fin_list:
			cell = chr(ascid+k) + str(value)
			k += 1
			#print cell, item
			ws[cell] = item
	# update_row_data END

#通过xml解析数组信息
def print_node(node):
	'''打印结点基本信息'''
	print "=============================================="
	print "node.attrib:%s" % node.attrib
	if node.attrib.has_key("id") is True:
		print "node.attrib['id']:%s" % node.attrib['id']
	print "node.tag:%s" % node.tag
	print "node.text:%s" % node.text

#解析XML文件，获取为每个个股定义的大单数组
def get_data_array(code, xmlfile):
	try:
		tree = ET.parse(xmlfile)
		root = tree.getroot()
	except Exception, e: 
		print "Parse file fail:", xmlfile
		return ""
	lst_node = root.getiterator("code")
	for node in lst_node:
		#print_node(node)
		if node.attrib.has_key("id") is False:
			continue
		id = node.attrib['id']
		if id!=code:
			continue
		#找出节点的子节点
		for child in node.getchildren():
			#print child.tag,':',child.text
			if child.tag=='array':
				return child.text
	return ""

def ts_handle_data(addcsv, prepath, bhist, url, code, qdate, replace, sarr):
	todayUrl = "http://hq.sinajs.cn/list=" + code
	#if Handle_Mid==0:
	#	print "Message: Ignore 中性盘"

	if not os.path.isdir(prepath):
		os.makedirs(prepath)

	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	#检查实时记录，是否出现大单
	check_vol1 = int(volObj[1])
	check_vol2 = int(volObj[3])
	Large_Volume = int(volObj[arrlen-1])*2

	totalline = 0
	filename = code+ '_' + qdate
	fctime = ''
	todayData = []
	todayDataLen = 0
	bGetToday = 0
	last_close = 0

	cur=datetime.datetime.now()
	if bhist==0:
		fctime = '%02d:%02d' %(cur.hour, cur.minute)
		filename = '%s_%02d-%02d' %(filename, cur.hour, cur.minute)
		bGetToday = 1
	else:
		#检查文件是否存在
		filexlsx = prepath +filename+ '.xlsx'
		if os.path.exists(filexlsx) and replace==0:
			#print "File Exist at:", filexlsx
			return -1
	filexlsx = prepath +filename+ '.xlsx'

	if bhist==2:
		bGetToday = 1
	if (bGetToday==1):
		try:
			req = urllib2.Request(todayUrl)
			stockData = urllib2.urlopen(req, timeout=5).read()
		except:
			loginfo(1)
			print todayUrl, " request timeout"
			return -1
		else:
			stockObj = stockData.split(',')
			if len(stockObj)<10:
				print code, ": No trade data"
				return -1

			closePrice = float(stockObj[3])
			lastClsPrice = float(stockObj[2])
			last_close = lastClsPrice
			openPrice = float(stockObj[1])
			highPrice = float(stockObj[4])
			lowPrice = float(stockObj[5])
			exVolume = int(stockObj[8])/100
			exAmount = float(stockObj[9])
			if closePrice==0.0 and highPrice==0.0:
				print code, ": Today TingPai ?"
				return -1
			f1 = '%02.02f'%( ((closePrice-lastClsPrice)/lastClsPrice)*100 )
			exFluc = float(f1)

			todayData.append(closePrice)
			todayData.append(exFluc)
			todayData.append(lastClsPrice)
			todayData.append(openPrice)
			todayData.append(highPrice)
			todayData.append(lowPrice)
			todayData.append(exVolume)
			todayData.append(exAmount)
			todayDataLen = len(todayData)

	if addcsv==1:
		filecsv = prepath + filename + '.csv'
		fcsv = open(filecsv, 'w')
		strline = '成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质'
		fcsv.write(strline)
		fcsv.write("\n")

	wb = Workbook()
	# grab the active worksheet
	ws = wb.active

	strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质,收盘价,涨跌幅,前收价,开盘价,最高价,最低价,成交量,成交额'
	strObj = strline.split(u',')
	ws.append(strObj)

	stockInfo = []
	#每一页的数据，如果找到匹配数据则设置为1；解决有时候页面有数据但是收不到，
	#count为0，重新加载尝试再次获取；如果解析到数据的页面，如果count为0就不再继续解析数据
	matchDataFlag = 0
	excecount = 0
	#保存9:30左右的大单和实时附近的大单
	savedTrasData = []
	savedTrasData2 = []
	largeTrasData = []
	i = 1

	curcode = code
	if len(code)==8:
		curcode = code[2:8]

	#历史当天行情数据通过此方法获得
	if bhist==1:
		df = ts.get_hist_data(curcode, start=qdate, end=qdate)
		if df is None or df.empty or len(df)!=1:
			if chk_holiday(qdate) is False:
				print qdate, ": No data"
			else:
				print qdate, "is holiday or weekend"
			return -1
		#print qdate, df
		for index,row in df.iterrows():
			last_close = float(row['close'])-float(row['price_change'])
			stockInfo.append(row['close'])
			stockInfo.append(row['p_change'])
			stockInfo.append(last_close)
			stockInfo.append(row['open'])
			stockInfo.append(row['high'])
			stockInfo.append(row['low'])
			stockInfo.append(row['volume'])
			#stockInfo.append(row['turnover'])
			stockInfo.append('')

	get_api = 0
	if bhist!=0:
		get_api = 1
	dtsrc = ['sn', 'tt', 'nt']
	dtidx = 0
	while excecount<=3:
		if get_api==0:
			df = ts.get_today_ticks(curcode)
		else:
			df = ts.get_tick_data(curcode, qdate, src=dtsrc[dtidx])
		if df is None:
			excecount += 1
			if excecount==4 and dtidx<3:
				excecount = 0
				dtidx += 1
			continue;
		if df.size==18:
			if bhist==2:
				get_api = 0
			excecount += 1
			print "Get data from source '%s' fail in %s" %(dtsrc[dtidx], qdate)
			dtidx += 1
			continue;
		else:
			break;

	#尝试3次，检查结果
	if df is None:
		print qdate, ": None Object"
		return -1
	if df.size==18:
		print qdate, ": Fail to get data"
		return -1

	lastTime=''
	lastvol = -1
	for index,row in df.iterrows():
		curtime = row['time']
		curvol = int(row['volume'])
		#检查2条数据是否重复，偶尔发现会有重复
		if lastTime!='':
			if lastTime==curtime:
				if lastvol==curvol:
					#print code, curtime, curvol, "discard repeat data!"
					continue
		pass
		lastTime = curtime
		lastvol = curvol
		curprice = row['price']
		range_per = ''
		if last_close!=0:
			range_val = ((float(curprice)-last_close) * 100) / last_close
			range_per = round(range_val, 2)
		fluctuate = row['change']
		volume = curvol
		amount = row['amount']
		if get_api==0:
			state = row['type']
		else:
			state = row['type'].decode('utf-8')

		ret,hour,minute,second = parseTime(curtime)
		if (ret==-1):
			continue
		if (int(amount)==0 and not (hour==15 and minute==0)):
			continue

		bAddVolumn = 1
		if (hour==9 and minute==25) or (hour==15 and minute==0):
			bAddVolumn = 0

		stateStr = state
		st_buy = '买盘'.decode('gbk')
		st_sell = '卖盘'.decode('gbk')
		st_mid = '中性盘'.decode('gbk')
		if cmp(state, st_sell)==0:
			if bAddVolumn==1:
				handle_volumn(volume, dataObj, 2)
			stateStr = 'SELL卖盘'.decode('gbk')
		elif cmp(state, st_buy)==0:
			if bAddVolumn==1:
				handle_volumn(volume, dataObj, 1)
		#目前中性盘没有处理
		elif cmp(state, st_mid)==0:
			if bAddVolumn==1:
				ret = handle_middle_volumn(volume, dataObj, curtime, fluctuate, 0)
			else:
				ret = 0
			if ret==1:
				stateStr = st_buy
			elif ret==2:
				stateStr = st_sell

		if addcsv==1:
			strline = curtime +","+ curprice +","+ range_per +","+ fluctuate +","+ curvol +","+ amount +","+ stateStr + "\n"
			fcsv.write(strline)

		totalline += 1
		row = totalline+1
		price = float(curprice)
		cell = 'A' + str(row)
		ws[cell] = curtime
		cell = 'B' + str(row)
		ws[cell] = price
		cell = 'C' + str(row)
		ws[cell] = range_per
		cell = 'D' + str(row)
		ftfluct = fluctuate
		if (fluctuate=='--'):
			ws[cell] = 0
		else:
			ftfluct = float(fluctuate)
			ws[cell] = ftfluct
		cell = 'E' + str(row)
		ws[cell] = curvol
		cell = 'F' + str(row)
		ws[cell] = int(amount)
		cell = 'G' + str(row)
		s1 = stateStr
		ws[cell] = s1

		if hour==9 and minute<30:
			flct = round(float(curprice)-last_close, 2)
			if fluctuate!=flct:
				cell = 'D' + str(row)
				ws[cell] = flct
			if bGetToday==1:
				if flct > 0:
					stateStr = '买盘'.decode('gbk')
				elif flct < 0:
					stateStr = 'SELL卖盘'.decode('gbk')
				else:
					stateStr = '中性盘'.decode('gbk')
				if state == '0':
					cell = 'G' + str(row)
					ws[cell] = stateStr

		#将开始和最后成交数据保存
		#实时的最新大单，保存在savedTrasData
		#9:30左右的大单，保存在savedTrasData1
		bSaveFlag = 0
		if (totalline==1 or (totalline<4 and curvol>check_vol1)):
			bSaveFlag = 1
		elif (hour==9 and minute==30 and curvol>check_vol2) or (hour==9 and minute<30):
			bSaveFlag = 2
		if bSaveFlag==1 or bSaveFlag==2:
			rowData = []
			rowData.append(curtime)
			rowData.append(price)
			rowData.append(range_per)
			rowData.append(ftfluct)
			rowData.append(curvol)
			rowData.append(int(amount))
			rowData.append(s1)
		if bSaveFlag==1:
			savedTrasData.append(rowData)
		elif bSaveFlag==2:
			savedTrasData2.append(rowData)

		#增加大单成交记录
		if (curvol>=Large_Volume):
			rowData = []
			rowData.append(curtime)
			rowData.append(price)
			rowData.append(range_per)
			rowData.append(ftfluct)
			rowData.append(curvol)
			rowData.append(int(amount))
			rowData.append(s1)
			largeTrasData.append(rowData)

	#将当天的数据在Sheet页面更新
	ascid = 72
	row2 = 2
	if ((bhist==0 or bhist==2) and todayDataLen>0):
		for k in range(0, todayDataLen):
			cell = chr(ascid+k) + str(row2)
			ws[cell] = todayData[k]
	elif bhist==1:
		number = len(stockInfo)
		for k in range(0,number):
			cell = chr(ascid+k) + str(row2)
			ws[cell] = stockInfo[k]

	ws.auto_filter.ref = "A1:G1"
	update_row_data(ws, filexlsx)

	if addcsv==1:
		fcsv.close()
		if (totalline==0):
			os.remove(filecsv)

	if totalline>0:
		startIdx = 0
		savedTrasLen = len(savedTrasData2)
		if savedTrasLen>0:
			if savedTrasLen>Tras_Count:
				startIdx = savedTrasLen-Tras_Count
			for j in range(startIdx, savedTrasLen):
				savedTrasData.append(savedTrasData2[j])
		ws = wb.create_sheet()
		write_statics(ws, fctime, dataObj, qdate, savedTrasData, largeTrasData)

	if (os.path.exists(filexlsx) and bhist==0):
		j = 1
		while True:
			filexlsx = prepath + filename + '_' + str(j) + '.xlsx'
			j += 1
			if not os.path.exists(filexlsx):
				break;

	wb.save(filexlsx)
	return 0

def ts_analyze_data(url, code, sarr, priceList, contPrice):
	url = url +"?symbol="+ code
	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	dataObjLen = len(dataObj)
	check_vol1 = int(volObj[1])
	check_vol2 = int(volObj[3])
	Large_Volume = int(volObj[arrlen-1])*2

	totalline = 0
	#可能数据在不同的页面，同时存在，这是重复数据需要过滤重复结果
	#还可能相同时间，产生多个成交量，需要都保留
	lasttime = ''
	lastvol = 0
	pageFtime = ''
	bFtime = 0
	hisUrl = ''
	fctime = ''
	todayData = []
	todayDataLen = 0
	bGetToday = 0

	cur=datetime.datetime.now()

	strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质,收盘价,涨跌幅,前收价,开盘价,最高价,最低价,成交量,成交额'
	strObj = strline.split(u',')
	#dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D')
	dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th>(.*)\D')
	frameRe = re.compile(r'.*name=\"list_frame\" src=\"(.*)\" frameborder')
	keyw = '收盘价|涨跌幅|前收价|开盘价|最高价|最低价|成交量|成交额'
	infoRe = re.compile(r'\D+('+keyw+').*>(\+?-?\d+\.\d+)')
	excecount = 0
	stockInfo = []
	reloadUrl = 0
	noDataFlag = 0
	noDataKey = "该股票没有交易数据"
	notTrasFlag = 0
	notTrasKey = "输入的日期为非交易日期"
	#每一页的数据，如果找到匹配数据则设置为1；解决有时候页面有数据但是收不到，
	#count为0，重新加载尝试再次获取；如果解析到数据的页面，如果count为0就不再继续解析数据
	matchDataFlag = 0
	i = 1
	lineCount = 0
	firstHour = 0
	firstMinute = 0
	firstSecond = 0
	minValue = 0
	maxValue = 0
	curValue = 0
	tmpContPrice = []
	pageIdx = 1
	bAlert = 0
	bLastAlert = 0

	st_buy = '买盘'.decode('gbk')
	st_sell = '卖盘'.decode('gbk')
	st_mid = '中性盘'.decode('gbk')

	#临时处理方案，对国债逆回购
	head5 = code[0:5]
	bBigChange = (cmp(head5, "sz131")==0) or (cmp(head5, "sh204")==0)

	#获取当天所有分笔交易
	curcode=code
	if len(code)==8:
		curcode=code[2:8]
	while excecount<=3:
		df = None
		try:
			df = ts.get_today_ticks(curcode)
		except:
			pass
		if df is None:
			excecount += 1
			continue
		excecount += 1
		if df.size>0:
			break

	print ""
	#尝试3次，检查结果
	if df is None:
		print curcode, ": None Object"
		return -1
	if df.size==0:
		print curcode, ": Fail to get today ticks"
		return -1

	for index,row in df.iterrows():
		curtime = row['time']
		curprice = row['price']
		range_per = ''
		#if last_close!=0:
		#	range_val = ((float(curprice)-last_close) * 100) / last_close
		#	range_per = round(range_val, 2)
		fluctuate = row['change']
		curvol = int(row['volume'])
		volume = curvol
		amount = row['amount']
		state = row['type']#.decode('utf8')
		lasttime = curtime
		lastvol = curvol
		volume = curvol
		#print curtime,curprice,row['pchange'],row['change'],curvol,amount,row['type'],state

		ret,hour,minute,second = parseTime(curtime)
		if (ret==-1):
			continue
		if (curprice=="0.00") or (fluctuate=="-100.00%"):
			print "page(%d) Price(%s) or range(%s) is invalid value"%(i, key.group(2), key.group(3))
			continue
		if (hour==9 and minute<=20) or (hour==15 and minute>1):
			count += 1
			continue

		if curvol>Large_Volume:
			#记录成交时间，判断防止数据重复
			bFind = 0
			for k in range(0, len(Large_Vol_Time)):
				if (curtime==Large_Vol_Time[k]):
					bFind=1
					break
			if bFind==0:
				Large_Vol_Time.append(curtime)
				sv = ''
				if state==st_sell:
					sv = 'S'
				elif state==st_buy:
					sv = 'B'
				elif state==st_mid:
					sv = 'M'

				#15分钟内的大单
				bMatch = time_range(firstHour, firstMinute, hour, minute, 15)
				if bMatch==1:
					if not bBigChange:
						msgstr = u'Hello Big_DT (%s	%s:%d)'%(curtime, sv, curvol)
						ctypes.windll.user32.MessageBoxW(0, msgstr, '', 0)

		if curvol>=check_vol2:
			if (len(tmpContPrice)==0):
				if (state==st_sell or state==st_buy):
					tmpContPrice.append(state)
					tmpContPrice.append(1)
					tmpContPrice.append(curvol)
			else:
				stateValue = tmpContPrice[0]
				allowAdd = tmpContPrice[1]
				if (allowAdd==1):
					if cmp(stateValue, state)==0:
						tmpContPrice.append(curvol)
					else:
						if state==st_sell:
							tmpContPrice[1] = 0
						elif state==st_buy:
							tmpContPrice[1] = 0
						elif state==st_mid:
							tmpContPrice.append(curvol)
			bAlert = handle_last_price(tmpContPrice, contPrice)
			if bLastAlert==1 and bAlert==0:
				print "Price=", contPrice
				totalVol = 0
				contPriceLen = len(contPrice)
				for k in range(0, contPriceLen):
					totalVol += contPrice[k]
				sv = 'BB'
				if tmpContPrice[0]==st_sell:
					sv = 'SS'
				if not bBigChange:
					msgstr = u'Continued(%s) %d: %d'%(sv, contPriceLen, totalVol/contPriceLen)
					ctypes.windll.user32.MessageBoxW(0, msgstr, '', 0)
				bLastAlert = bAlert
			elif bAlert==1 and bLastAlert==0:
				bLastAlert = bAlert

	priceList[0] = curValue
	priceList[1] = int(curValue*100)
	priceList[2] = minValue
	priceList[3] = maxValue
	return 0

def list_stock_news(code, curdate, file):
	df = None
	try:
		df = ts.get_notices(code, curdate)
	except:
		pass

	if df is None or df.empty:
		df = ts.get_notices(code)
	for index,row in df.iterrows():
		if index>1:
			break
		print row['date'],row['title']
	print ''

def list_latest_news(codeArray, curdate):
	cur_dt = datetime.datetime.strptime(curdate, '%Y-%m-%d').date()
	for code in codeArray:
		bFlag = 0
		news_ct = 0
		df = ts.get_notices(code)
		for index,row in df.iterrows():
			news_dt = datetime.datetime.strptime(row['date'], '%Y-%m-%d').date()
			delta = cur_dt-news_dt
			if delta.days<=0:
				print "==================================="
				bFlag=1
			else:
				if bFlag==0:
					news_ct += 1
				else:
					break
				if news_ct>1:
					break
			print row['date'],row['title']
		print ''

def show_index_info(df, show_idx):
	if df is None:
		return
	for index,row in df.iterrows():
		if row[0] not in show_idx:
			continue

		open = float(row['open'])
		close = float(row['close'])
		preclose = float(row['preclose'])
		print "%8.2f(%6s)"%(close, row[2])

def show_extra_index(codeArray):
	df = ts.get_realtime_quotes(codeArray)
	for index,row in df.iterrows():
		pre_close = row['pre_close']
		price = row['price']
		f_pclose = float(pre_close)
		f_price = float(price)
		value =  (f_price-f_pclose)*100/f_pclose
		print "%8.2f( %3.2f)"%(f_price, round(value,2))

def check_cx(code):
	if len(code)!=6:
		return 0
	if code.isdigit() is False:
		return 0

	df = ts.get_hist_data(code)
	if df is None:
		return 0
	if len(df)>39:
		return 0

	b_match = 1
	for index,row in df.iterrows():
		open = float(row['open'])
		close = float(row['close'])
		high = float(row['open'])
		low = float(row['close'])
		p_change = float(row['p_change'])
		if p_change>12:
			continue
		if open==close and low==high and p_change>1:
			pass
		else:
			b_match = 0
			break
	return b_match

#分析时间得到ZT时间，上午or下午
def zt_time_analyze(chuban, stcsItem):
	timeObj = re.match(r'(\d{2}):(\d{2})', chuban)
	hour = int(timeObj.group(1))
	minute = int(timeObj.group(2))
	if hour<=11:
		stcsItem.s_sw_zt += 1
	else:
		stcsItem.s_xw_zt += 1

#通过分时得到ZT时间，上午or下午
def zt_time_point(code, zt_price, trade_date, stcsItem):
	excecount = 0
	df = None
	while excecount<=3:
		try:
			df = ts.get_k_data(code, ktype='60')
		except:
			excecount += 1
		else:
			break
	if df is None:
		return

	df_today = df.loc[df['date'].str.contains(str(trade_date))]
	if len(df_today)<=0:
		#print "Not find data"
		return
	j=1
	#读出的数据就是float类型
	#通过定义区间，记录ZT的时间段
	#1:[9:30-10:30],2:[10:30-11:30]...
	for index,row in df_today.iterrows():
		high = row['high']
		if high==zt_price:
			timeObj = re.match(r'.* (\d{2}):(\d{2})', row['date'])
			if (timeObj is None):
				print code, "非法时间格式1：" +str(row['date'])+ ", 期望格式: HH:MM"
				continue
			hour = int(timeObj.group(1))
			minute = int(timeObj.group(2))
			if hour<=11:
				stcsItem.s_sw_zt += 1
			else:
				stcsItem.s_xw_zt += 1
			break
		else:
			j += 1
	return

def get_zf_days1(code, type, trade_date, cur_zdt=1):
	print code, trade_date
	excecount=0
	df = None
	while excecount<=3:
		try:
			df = ts.get_hist_data(code)
		except:
			excecount += 1
		else:
			break
	if df is None:
		return 0
	#如果当天的数据已经得到，不需要多加一天，得到ZT or DT
	#只有正在进行交易的时候，当天数据得不到，ZT or DT需要加1
	count = 0
	for index,row in df.iterrows():
		val = float(row['p_change'])
		bflag = 0
		if str(index)==str(trade_date):
			if cur_zdt==1:
				count += 1
			continue
		if type==1:
			if val>9.9:
				count += 1
				bflag = 1
		elif type==2:
			if val<-9.88:
				count += 1
				bflag = 1
		if bflag==0:
			break
	pass
	return count

urlfmt = 'http://pdfm2.eastmoney.com/EM_UBG_PDTI_Fast/api/js?id=%s&TYPE=k&js=%s((x))&rtntype=%d&isCR=true&authorityType=fa'
#'Referer': http://quote.eastmoney.com/chart/h5.html?id=0005202&type=k
send_headers = {
'Host': 'pdfm2.eastmoney.com',
#'Connection': 'keep-alive',
'Cache-Control': 'max-age=0',
#'Upgrade-Insecure-Requests': 1,
'User-Agent':'Mozilla/5.0 (Windows NT 6.3; WOW64; rv:53.0) Gecko/20100101 Firefox/53.0',
'Accept': '*/*',
'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
'Accept-Encoding': 'gzip, deflate',
'Cookie': 'st_pvi=65131814925934; em_hq_fls=js; emhq_stock=300379%2C510300%2C300059; \
ct=gSzyZYX428bqF8KFmsvff_vR0WGKJAA0miUW7eR26gofMoK0iklrNPydT7ek5CAYmhkuwM3wbZOI2ZpFPAoWJelbJErRFVm7cCuHKYCtEJprcFnO6Q8wbTA_3JtbNlgYgCT2__I5z8tzKkfVANJw1xeOMTqksUHhEHm5G40vJ_E \
ut=FobyicMgeV7CWBTmL9Bu1obPhorDA1HbLw1lQWu36v9QOPj992o3gmSvzNX--G0lXhxZqa2MTa_B9Do2KSntZC_ssXb_gCrvSlax_wBf65973PgGLaxGqrqIXmGZFpMqsWk0SUMlQOCNdAg5wVaGd2GzdasuwK_fdFGtI9iJzvtEFnDsNVI5TGTtSqkoYM-oSZn1cY-cKZKPFHub6Gy75x68Dx1tTm5h7LywruhRDToUGB1laoHZIAJbVfqqyUI1RfICpFEaILg5jMTa9eNJPTOGX1YmEYHM; \
pi=6100112247957528%3byudingding6197%3bgoutou%3b0cW4DqnbVJoMdu1MZiropTm6BV1XLj16XsvJO70kiLEUyKsPMgichDgFLNQAsInwhJERWZz7yHcS7O7W3ZL2dcjA3%2fmZ3C5963s%2bDmkHccsL6pryylhnTIMnse5nSBAA9u4HViDawKBE2J1H99l9hN9FV2EDvHkoUmqLb6Yyf09WAifb%2bdYj9kNpyqiJziuNO7OTlZ8N%3bvNvIvbNvs29ty0dGWHDrlmY9DQZgV43TfP00Z0d5NjJm2tRrZqJ0843uLtzmM0c1G%2fs%2f3JarUusrH2NZdgcxiqLEIitAnfBpEb68QjYndEfZINJ%2bM1Pkry%2fnzvKD6XsX7nSGNF5EmGXIR23aUOrXivsh0bSqVQ%3d%3d; \
uidal=6100112247957528goutou; sid=4725419; vtpst=|; qgqp_b_id=3351a1846039a4841a7ac4edce9be185; st_si=57858667318678; \
HAList=a-sz-300059-%u4E1C%u65B9%u8D22%u5BCC%2Ca-sz-002252-%u4E0A%u6D77%u83B1%u58EB%2Ca-sz-300401-%u82B1%u56ED%u751F%u7269%2Ca-sh-600908-%u65E0%u9521%u94F6%u884C%2Ca-sz-002527-%u65B0%u65F6%u8FBE%2Ca-sh-600696-ST%u5339%u51F8%2Ca-sh-603618-%u676D%u7535%u80A1%u4EFD%2Ca-sz-300291-%u534E%u5F55%u767E%u7EB3%2Ca-sz-002815-%u5D07%u8FBE%u6280%u672F%2Ca-sh-603156-%u517B%u5143%u996E%u54C1%2Ca-sh-600845-%u5B9D%u4FE1%u8F6F%u4EF6; \
st_sn=1; st_psi=20180421183808157-117005300001-2337497230; st_asi=delete; EMSTtokenId=3bbbf98339c2b81dc9c8a173b9a51c54',
'DNT': 1
}
'''
'Cookie': 'st_pvi=97140791819816; emstat_bc_emcount=24578815992596371324; emstat_ss_emcount=0_1515626347_1914456706; \
em_hq_fls=old; _ga=GA1.2.1371087654.1472141391; Hm_lvt_557fb74c38569c2da66471446bbaea3f=1506330203; qgqp_b_id=867a67ca83c13fbd622d079314b267df; \
ct=tTNAoSxXA4HgHmLlK58m5dvCXLfX46zX2EhPxEC0Mp9KXne6YtatYD6Zc0kmqHFH9fEfxnQyVM3cy4d5hDSq4xaCuyjCtoat4MAwVJeFCgQ4jH3xU77tzv2BnAs_YtDRJ5YbPVra0XcbM0-zTxO2seQHPw9FjR4vaKjkGFEoupI; \
uidal=6100112247957528goutou; vtpst=|; emshistory=%5B%22%E4%BF%9D%E5%8D%83%E9%87%8C%E5%80%BA%E5%88%B8%22%5D; st_si=56676665040049',
'''
#返回值是ZDT的天数，stk_list[0]表示上市交易的天数，判断是否是次新
def get_zf_days(code, type, trade_date, cur_zdt, stk_list):
	rtntype = 1
	jstr = 'fsData1515847425760'
	excecount=0
	df = None

	ret, ncode = parseCode(code, 'dc')
	if ret!=0:
		exit(-1);

	urlall = urlfmt %(ncode, jstr, rtntype)
	#print("get_zf_days", urlall)
	while excecount<=5:
		try:
			req = urllib2.Request(urlall,headers=send_headers)
			res_data = urllib2.urlopen(req, timeout=3)
		except:
			excecount += 1
			continue
		else:
			break
	if res_data is None:
		print("Open URL fail", sys._getframe().f_code.co_name,code, urlall)
		return -1

	content1 = res_data.read()
	respInfo = res_data.info()
	if( ("Content-Encoding" in respInfo) and (respInfo['Content-Encoding'] == "gzip")):
		content1 = zlib.decompress(content1, 16+zlib.MAX_WBITS);
	else:
		print "Content not zip"
	res_data.close()

	content = content1.decode('utf8')
	if content[-1]==")":
		content = content[:-1]
	jslen = len(jstr)
	content = content[(jslen+1):]
	contObj = content.split("\r")
	
	#如果当天的数据已经得到，不需要多加一天，得到ZT or DT
	#只有正在进行交易的时候，当天数据得不到，ZT or DT需要加1
	count = 0
	yzcount = 0
	dayLen = len(contObj)
	stk_list[0] = dayLen
	while dayLen>0:
		dayCont = contObj[dayLen-1]
		if len(dayCont)<=8:
			dayLen -= 1
			continue
		#print(dayCont)
		itemObj = dayCont.strip().split(',')

		index = itemObj[0]
		close = float(itemObj[2])
		high = float(itemObj[3])
		low = float(itemObj[4])
		#最开始上市那天，以前不再有数据
		if dayLen<=1:
			count += 1
			yzcount += 1
			if yzcount!=0:
				count -= yzcount
			break
		preDayCont = contObj[dayLen-2]
		preItemObj = preDayCont.strip().split(',')
		preClose = float(preItemObj[2])

		val = round((close-preClose)*100/preClose, 2)
		bflag = 0
		if str(index)==str(trade_date):
			if cur_zdt==1:
				count += 1
			dayLen -= 1
			continue
		#print index, trade_date, type, val
		if type==1:
			if (val>9.8 and high==close) or (high==low and val>2):
				count += 1
				if high==low:
					yzcount += 1
				bflag = 1
		elif type==2:
			if (val<-9.88 and low==close) or (high==low and val<-2):
				count += 1
				if high==low:
					yzcount += 1
				bflag = 1
		if bflag==0:
			break
		dayLen -= 1
	return count

def handle_today_ticks(df, code, trade_date, chk_price, type):
	tmstr = '??:??'
	return tmstr

# minute: 期望减去几分钟，如果是涨跌停开板时间不用减
# flag: 是否加上？号，有的涨跌停就是一瞬间，意义不大
def covert_time_fmt(tmobj, minute, flag):
	#将当前时间减去5分钟
	dt = datetime.datetime.strptime(tmobj, "%Y-%m-%d %H:%M")
	newdt = dt - datetime.timedelta(minutes=minute)
	tmobj = newdt.strftime("%Y-%m-%d %H:%M")

	timeObj = re.match(r'.* (\d{2}):(\d{2})', tmobj)
	if (timeObj is None):
		print code, "非法时间格式2：" +str(row['date'])+ ", 期望格式: HH:MM"
		return ''
	hour = int(timeObj.group(1))
	minute = int(timeObj.group(2))
	if hour<=11:
		if flag==1:
			tmstr = "%02d:%02d??" %(hour, minute)
		else:
			tmstr = "%02d:%02d" %(hour, minute)
	else:
		if flag==1:
			tmstr = "%02d:%02d--??" %(hour, minute)
		else:
			tmstr = "%02d:%02d--" %(hour, minute)
	#if hour<=11:
	#	tmstr = "%02d:%02d??" %(hour, minute)
	#else:
	#	tmstr = "%02d:%02d--??" %(hour, minute)
	return tmstr

#type  0:ZT  1:DT
def handle_kdata(df, code, trade_date, chk_price, type, tm_array):
	tmstr = '??:??'
	df_today = df.loc[df['date'].str.contains(str(trade_date))]
	if len(df_today)<=0:
		print code, trade_date, "No data"
		#print "Not find data"
		return

	tmobj = ''
	tmend = ''
	mx_prc = 0
	#读出的数据就是float类型
	for index,row in df_today.iterrows():
		close = row['close']
		if type==0:
			price = row['high']
			if mx_prc<price:
				mx_prc = price
				tmobj = row['date']
				tmend = row['date']
			elif mx_prc==price:
				tmend = row['date']
		else:
			price = row['low']
			if mx_prc==0:
				mx_prc = price
				tmobj = row['date']
				tmend = row['date']
			if mx_prc>price:
				mx_prc = price
				tmobj = row['date']
				tmend = row['date']
			elif mx_prc==price:
				tmend = row['date']

	binst = 0
	if mx_prc!=chk_price:
		binst = 1

	if tmobj!='':
		tmstr = covert_time_fmt(tmobj, 5, binst)
		tm_array.append(tmstr)
	if tmend!='':
		tmstr = covert_time_fmt(tmend, 0, binst)
		tm_array.append(tmstr)
	return tmstr

#获取首次触板ZT or DT的时间
#type  0:ZT  1:DT
def get_zdt_time(code, trade_date, chk_price, type, tm_array):
	tmstr = '??:??'
	excecount=0
	df = None
	flag = 0
	today = datetime.date.today()
	if (today-trade_date).days>0:
		flag = 1
	#修改一下这里，是否k_data就可以得到当天K线了
	flag = 1
	while excecount<=3:
		try:
			if flag == 0:
				df = ts.get_today_ticks(code)
			else:
				df = ts.get_k_data(code, ktype='5')
		except:
			excecount += 1
		else:
			break
	if df is None:
		return tmstr

	if flag == 0:
		tmstr = handle_today_ticks(df, code, trade_date, chk_price, type)
	else:
		tmstr = handle_kdata(df, code, trade_date, chk_price, type, tm_array)
	#print code, type, tmstr
	return tmstr

#分析此单的状态：ZT、DT、ZTHL、DTFT...，详见 statisticsItem 定义
def analyze_status(code, name, row, stcsItem, yzcx_flag, pd_list, trade_date):
	if len(code)!=6:
		return 0
	if code.isdigit() is False:
		return 0

	status = 0
	high = round(float(row['high']),2)
	low = round(float(row['low']),2)
	open = round(float(row['open']),2)
	price = round(float(row['price']),2)
	pre_close = round(float(row['pre_close']),2)
	volume = int(row['volume'])

	#排除当天没有交易
	if pre_close==0:
		print "PRE_CLOSE Not exist!!!", code, name
		return 0

	if high==0 or low==0:
		return 0

	b_ST = 0
	if name.find("ST")>=0 or name[0:1]=="S":
		b_ST = 1
		#print code,name

	o_percent = (open-pre_close)*100/pre_close
	c_percent = (price-pre_close)*100/pre_close
	h_percent = (high-pre_close)*100/pre_close
	l_percent = (low-pre_close)*100/pre_close
	open_percent = spc_round(o_percent,2)
	change_percent = spc_round(c_percent,2)
	high_zf_percent = spc_round(h_percent,2)
	low_df_percent = spc_round(l_percent,2)

	#获取ZT
	if b_ST==1:
		zt_price1 = pre_close * 1.05
		dt_price1 = pre_close * 0.95
	else:
		zt_price1 = pre_close * 1.1
		dt_price1 = pre_close * 0.9
	zt_price = spc_round(zt_price1,2)
	dt_price = spc_round(dt_price1,2)
	#print name, pre_close, zt_price, dt_price

	#YZ状态处理
	stk_list = [0, 0]
	if high==low:
		if open>pre_close:
			#刚开盘，只有1笔成交，会大量的输出，不合适检查
			#if high!=zt_price:
			#	print "Warning: YZ, not ZT??? ", code, high, zt_price
			if b_ST==1 and high==zt_price:
				stcsItem.s_st_yzzt += 1
				status |= STK_ST_YZZT
			elif b_ST==0 and high==zt_price:
				if high_zf_percent>15:
					pass
				else:
					stcsItem.s_yzzt += 1
					status |= STK_YZZT
					stcsItem.s_zt += 1
					status |= STK_ZT
					stcsItem.s_open_zt += 1
					status |= STK_OPEN_ZT
					stcsItem.s_close_zt += 1
					#list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent]
					#pd_list.append(list)
					if yzcx_flag==0:
						count = get_zf_days(code, 1, trade_date, 1, stk_list)
						if stk_list[0]<CX_DAYS:
							stcsItem.s_cxzt += 1
						list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent, count, stk_list[0]]
						stcsItem.lst_non_yzcx_yzzt.append(list)
					#print stcsItem.s_zt,code,name,price,change_percent,open
		elif open<pre_close:
			#if low!=dt_price:
			#	print "Warning: YZ, not DT??? ", code, low, dt_price
			if b_ST==1 and low==dt_price:
				stcsItem.s_st_yzdt += 1
				status |= STK_ST_YZDT
			elif b_ST==0 and low==dt_price:
				stcsItem.s_yzdt += 1
				status |= STK_YZDT
				stcsItem.s_dt += 1
				status |= STK_DT
				stcsItem.s_open_dt += 1
				status |= STK_OPEN_DT
				count = get_zf_days(code, 2, trade_date, 1, stk_list)
				if stk_list[0]<CX_DAYS:
					stcsItem.s_cxdt += 1
				list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent, count, stk_list[0]]
				stcsItem.lst_yzdt.append(list)
		#print code,name,open,low,high,price,pre_close
	else:
		if high==zt_price:
			if b_ST==0:
				tmArr = []
				get_zdt_time(code, trade_date, zt_price, 0, tmArr)
				chuban = tmArr[0]
				openban = tmArr[1]
				zt_st = ''
				if price==zt_price:
					#仅仅计算最后还是ZT的item
					#zt_time_point(code, zt_price, trade_date, stcsItem)
					zt_time_analyze(chuban, stcsItem)
					stcsItem.s_zt += 1
					status |= STK_ZT
					if open==zt_price:
						stcsItem.s_open_zt += 1
						status |= STK_OPEN_ZT
						if price==open:
							stcsItem.s_close_zt += 1
							stcsItem.s_open_T_zt += 1
							zt_st = 'T'
					#list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent]
					#pd_list.append(list)
					count = get_zf_days(code, 1, trade_date, 1, stk_list)
					if yzcx_flag==0:
						if stk_list[0]<CX_DAYS:
							stcsItem.s_cxzt += 1
						list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent, count, stk_list[0], chuban, zt_st]
						stcsItem.lst_non_yzcx_zt.append(list)
				else:
					count = get_zf_days(code, 1, trade_date, 0, stk_list)
					stcsItem.s_zthl += 1
					status |= STK_ZTHL
					if open==zt_price:
						stcsItem.s_open_zt += 1
						status |= STK_OPEN_ZT
						stcsItem.s_dk_zt += 1
						zt_st = 'K'
						#print stcsItem.s_zthl,code,name,price,high,zt_price
					list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent, count, stk_list[0], chuban, openban, zt_st]
					stcsItem.lst_non_yzcx_zthl.append(list)
				if change_percent<=3:
					stcsItem.lst_kd.append(name)
				if price<open:
					stcsItem.s_zt_o_gt_c += 1

		if low==dt_price:
			if b_ST==0:
				tmArr = []
				get_zdt_time(code, trade_date, dt_price, 1, tmArr)
				chuban = tmArr[0]
				openban = tmArr[1]
				dt_st = ''
				if open==dt_price:
					stcsItem.s_open_dt += 1
					status |= STK_OPEN_DT
					if price!=dt_price:
						stcsItem.s_open_dt_dk += 1

				if price==dt_price:
					stcsItem.s_dt += 1
					status |= STK_DT
					if open==dt_price:
						dt_st = 'DT'
					count = get_zf_days(code, 2, trade_date, 1, stk_list)
					if stk_list[0]<CX_DAYS:
						stcsItem.s_cxdt += 1

					#DT Data
					list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent, count, stk_list[0], chuban, dt_st]
					stcsItem.lst_dt.append(list)
				else:
					stcsItem.s_dtft += 1
					status |= STK_DTFT
					if open==dt_price:
						dt_st = 'K'

					#DTFT Data
					count = get_zf_days(code, 2, trade_date, 0, stk_list)
					list = [code, name, change_percent, price, open_percent, high_zf_percent, low_df_percent, count, stk_list[0], chuban, openban, dt_st]
					stcsItem.lst_dtft.append(list)

	#统计开盘涨跌幅度
	if open_percent>0:
		#print("%s,%s,%f"%(code,name,open_percent))
		stcsItem.s_open_sz += 1
		if open_percent>=4.0:
			stcsItem.s_open_dz += 1
	elif open_percent<0:
		stcsItem.s_open_xd += 1
		if open_percent<=-4.0:
			stcsItem.s_open_dd += 1
	else:
		stcsItem.s_open_pp += 1

	#统计开盘涨跌幅度
	if change_percent>0:
		stcsItem.s_close_sz += 1
		if change_percent>=4.0:
			stcsItem.s_close_dz += 1
	elif change_percent<0:
		stcsItem.s_close_xd += 1
		if change_percent<=-4.0:
			stcsItem.s_close_dd += 1
	else:
		stcsItem.s_close_pp += 1
	stcsItem.s_total += 1

	#统计最大涨跌幅度
	if high_zf_percent>=4.0:
		stcsItem.s_high_zf += 1
	if low_df_percent<=-4.0:
		stcsItem.s_low_df += 1

	#统计表现震撼个股 排除YZZT的新股
	zf_range = high_zf_percent-low_df_percent
	if zf_range>=15.0 and yzcx_flag==0:
		if change_percent>=6.0:
			list = [code, name, change_percent, price, zf_range]
			stcsItem.lst_nb.append(list)
		elif change_percent<=-6.0:
			list = [code, name, change_percent, price, zf_range]
			stcsItem.lst_jc.append(list)
	return status

#建议在之前调用debug\instant_data.py更新所有数据，在data\entry\trade\trade_last.txt保存
def get_guben_line(url_str, code):
	url = url_str % (code)
	#print url
	guben_info = None
	excecount = 0
	while excecount<=3:
		try:
			req = urllib2.Request(url)
			guben_info = urllib2.urlopen(req, timeout=5).read()
		except:
			excecount += 1
		else:
			break
	if guben_info is None:
		print "Fail to get guben data"
		return None
	#找到匹配数据
	start = guben_info.find("<set name='")
	if start<0:
		return None
	end = guben_info[start:].find("</graph>")
	if end<0:
		return None
	return guben_info[start:start+end]

def parse_guben(gb_str, gb_list):
	while gb_str:
		rec = []
		obj = re.match(r'<set name=\'.*?\' value=\'(.*?)\' hoverText=\'(.*?)\'/>(.*)', gb_str)
		#print obj,obj.group(1),obj.group(2)
		rec.append(obj.group(2))
		if obj.group(1)=='':
			rec.append(0)
		else:
			rec.append(float(obj.group(1)))
		gb_list.append(rec)
		gb_str = obj.group(3)
	return

#不通过get_today_all()接口，使用东财接口
def get_today_new_stock(new_st_list):
	LOOP_COUNT=0
	'''
	st_today_base = None
	while LOOP_COUNT<3:
		try:
			st_today_base = ts.get_today_all()
		except:
			LOOP_COUNT += 1
			time.sleep(0.5)
		else:
			break
	if st_today_base is None:
		print "Timeout to get stock basic info, check new stk info manually!!!"
	else:
		st_today_df = st_today_base.sort_values(['changepercent'], 0, False)
		for index,row in st_today_df.iterrows():
			code = row[0].encode('gbk')
			if row['changepercent']>11:
				new_st_list.append(code)
	print ''
	'''
	
	send_headers = {
	'Host': 'nufm.dfcfw.com',
	'DNT': 1,
	'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.101 Safari/537.36',
	'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
	'Accept-Encoding': 'gzip, deflate',
	'Accept-Language': 'zh-CN,zh;q=0.8'
	}
	url = "http://dcfm.eastmoney.com/EM_MutiSvcExpandInterface/api/js/get?callback=jQuery1123019195110268527693_1616377243676&st=purchasedate%2Csecuritycode&sr=-1&ps=50&p=1&type=XGSG_LB&js=%7B%22data%22%3A(x)%2C%22pages%22%3A(tp)%7D&token=894050c76af8597a853f5b408b759f5d	LOOP_COUNT = 0"
	url = "http://nufm.dfcfw.com/EM_Finance2014NumericApplication/JS.aspx?type=CT&cmd=C._A&sty=FCOIATA&sortType=C&sortRule=-1&page=1&pageSize=20&js={rank:[(x)],pages:(pc)}&token=7bc05d0d4c3c22ef9fca8c2a912d779c"
	response = None
	#print url
	while LOOP_COUNT<3:
		try:
			req = urllib2.Request(url, headers=send_headers)
			response = urllib2.urlopen(req, timeout=5)
		except:
			LOOP_COUNT += 1
		else:
			break
	if response is None:
		print "Please check no data from DongCai"
		return

	content = response.read()
	respInfo = response.info()
	line = content
	if( ("Content-Encoding" in respInfo) and (respInfo['Content-Encoding'] == "gzip")):
		#print "Content compressed"
		line = zlib.decompress(content, 16+zlib.MAX_WBITS);

	print("line", line)
	obj = re.match(r'{rank:\["(.*)"\].*', line)
	rank = obj.group(1)
	array = rank.split('","')
	for i in range(0, len(array)):
		props = array[i].split(',')
		if props[2][0]=='N':
			code = props[1]
			new_st_list.append(code)
	return

#TODO: 以下几个函数可以合并
def get_trade_date(zsidx='sh'):
	df = ts.get_k_data(zsidx)
	if df is None:
		print "ts get k data fail"
		return None
	list = df['date'].tolist()
	ll = sorted(list, reverse=True)
	return ll

def get_last_trade_dt(zsidx='sh'):
	list = get_trade_date(zsidx)
	if len(list)>0:
		return list[0]
	return None

def get_pre_trade_date(tradeList, days, code='399001'):
	today = datetime.date.today()
	td_str = '%d-%02d-%02d' %(today.year, today.month, today.day)
	df = ts.get_hist_data(code)
	if df is None:
		return
	count = 0
	#idxlist=list(df.index)
	for index,row in df.iterrows():
		if count>=days:
			break
		#if td_str==index:
		#	continue
		tradeList.append(index)
		count += 1

def get_all_stk_info(st_list, dc_data, today_open, stcsItem):
	sysstr = platform.system()
	today = datetime.date.today()
	number = len(st_list)
	if number<=0:
		print("st_list is NULL")
		return -1

	#得到ZS的信息，但是当天交易的时候，得不到当天的
	delta1=datetime.timedelta(days=30)
	sdate = today-delta1
	fmt_start = '%d-%02d-%02d' %(sdate.year, sdate.month, sdate.day)
	kdf = ts.get_k_data('000001', index=True, start=fmt_start)
	kdf = kdf.sort_values(['date'], 0, False)
	last_idx_date = kdf.iloc[0,0]
	#idx_date 得到最近一天交易的日期
	idx_date = datetime.datetime.strptime(last_idx_date, '%Y-%m-%d').date()

	#得到最近一天的实时信息，得到日期
	rq_idx = get_last_trade_dt()
	print rq_idx
	rq_idx_dt = datetime.datetime.strptime(rq_idx, '%Y-%m-%d').date()

	cmp_delta = rq_idx_dt-idx_date
	if cmp_delta.days>0:
		idx_date = rq_idx_dt

	b_get_data = 1
	#ZT一次取出 base 个
	#截取list，通过配置起始位置
	base = 23
	loop_ct = number/base
	if number%base!=0:
		loop_ct += 1

	pd_list = []
	for i in range(0, loop_ct):
		end_idx = min(base*(i+1), number)
		cur_list = st_list[i*base:end_idx]
		if len(cur_list)==0:
			break
		#print cur_list
		LOOP_COUNT = 0
		stdf = None
		while LOOP_COUNT<5:
			try:
				stdf = ts.get_realtime_quotes(cur_list)
			except:
				time.sleep(0.5)
				LOOP_COUNT += 1
			else:
				break
		if stdf is None:
			print "Get list fail at:", cur_list
			return -1

		#print stdf
		for index,row in stdf.iterrows():
			stockInfo = []
			code = cur_list[index]
			# Ignore KCB TODO: support it
			if (code[0:3]=='688'):
				#print("Filter " + code)
				continue
			index += 1
			if sysstr=="Windows":
				name = row[0].encode('gbk')
			elif sysstr == "Linux":
				name = row[0].encode('utf8')
			pre_close = float(row['pre_close'])
			price = float(row['price'])
			volumn = int(row['volume'])

			ask = float(row['ask'])
			if volumn==0 and dc_data==1:
				return 0
			if pre_close==0:
				print ("%s %s invalid value" %(code,name))
				continue
			change_perc = (price-pre_close)*100/pre_close
			today_high = float(row['high'])
			today_low = float(row['low'])
			today_b1_p = float(row['b1_p'])
			today_a1_p = float(row['a1_p'])
			today_bid = float(row['bid'])
			#判断今日是否trade suspend
			if today_high==today_low and today_high==0.0:
				continue
			elif today_b1_p==0.0 and today_a1_p==0.0 and today_bid==0.0:
				continue

			#通过获得K线数据，判断是否YZZT新股
			yzcx_flag = 0
			if b_get_data == 1:
				#获得每只个股每天交易数据
				day_info_df = None
				#新股上市可能有Bug
				try:
					day_info_df = ts.get_k_data(code)
				except:
					print "Error for code:", code, name
				if day_info_df is None:
					continue
				#print code, day_info_df
				trade_days = len(day_info_df)

				b_open=0
				yzzt_day = 0
				if trade_days==0:
					if name[0:1]=='N':
						stcsItem.s_new += 1
						yzcx_flag = 1
						continue
				if trade_days==1:
					if name[0:1]=='N':
						stcsItem.s_new += 1
						yzcx_flag = 1
				for tdidx,tdrow in day_info_df.iterrows():
					open = tdrow[1]
					close = tdrow[2]
					high = tdrow['high']
					low = tdrow['low']
					if high!=low:
						if yzzt_day!=0:
							if (yzzt_day+1)==trade_days:
								chg_perc = round((price-pre_close)*100/pre_close,2)
								open_list = [code, name, chg_perc, price, yzzt_day]
								today_open.append(open_list)
							b_open = 1
							break
					#当ZT打开，就会break for 循环
					yzzt_day += 1
					pre_close = close
				if b_open==0:
					try:
						dt_str=day_info_df.iloc[trade_days-1,0]
					except:
						print code,trade_days
						print day_info_df
						exit()
					last_date = datetime.datetime.strptime(dt_str, '%Y-%m-%d').date()
					#print code, name, idx_date,last_date
					cmp_delta = idx_date-last_date
					if cmp_delta.days==0:
						#print("%s,%s,CX YZZT"%(code,name))
						stcsItem.s_cx_yzzt += 1
						yzcx_flag = 1

				#如果是通过东财获得个股排行，不要判断新股的上市天数
				if dc_data==0:
					#认为YZZT不会超过 33 个交易日
					if trade_days>33:
						b_get_data = 0
				else:
					if trade_days>1:
						l_volume = float(day_info_df.iloc[trade_days-2,2])
						l_pclose = float(day_info_df.iloc[trade_days-2,2])
						l_close = float(day_info_df.iloc[trade_days-1,2])
						l_high = float(day_info_df.iloc[trade_days-1,3])
						chg_perc = round((l_close-l_pclose)*100/l_pclose,2)
						if l_close != l_high or chg_perc<9.5:
							b_get_data = 0
			# end if b_get_data
			#print(index)
			stk_type = analyze_status(code, name, row, stcsItem, yzcx_flag, pd_list, idx_date)
	return 0
