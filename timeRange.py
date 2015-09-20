# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from internal.common import *

prepath = '../Data/'
allFileNum = 0
dftArray = [0,200,300,600,900]
g_sttime = ''
g_edtime = ''

def parseFile(wkfile, fltarray, stdelta, eddelta):
	wb = load_workbook(wkfile)
	if "Sheet" not in wb.get_sheet_names():
		return
	ws = wb.get_sheet_by_name(name='Sheet')
	max_column = ws.max_column
	if max_column<7:
		print "column(%d) too short, please check" % max_column
		return

	dataObj = []
	arrlen = len(fltarray)
	for i in range(0,arrlen):
		obj = fitItem(fltarray[i])
		dataObj.append(obj)

	for rx in range(1,ws.max_row):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value
		w7 = ws.cell(row = rx, column = 7).value
		if (w1 is None) and (w2 is None) and (w3 is None) and (w4 is None)\
			and (w5 is None) and (w6 is None) and (w7 is None):
			break
		if (w1 is None) or (w2 is None) or (w5 is None) or (w6 is None) or (w7 is None):
			print "某项记录为None，不正确", w1, w2, w5, w6, w7 
			continue
		if rx==1:
			continue

		state = w7
		for j in range(0, arrlen):
			if w5 < fltarray[j]:
				continue
			timeObj = re.match(r'(\d+):(\d+):(\d+)', w1)
			if timeObj is None:
				print "Invalid time:", w1
				continue
			hour = int(timeObj.group(1))
			minute = int(timeObj.group(2))
			second = int(timeObj.group(3))
			timedlt = datetime.timedelta(hours=hour, minutes=minute, seconds=second).seconds
			if timedlt<stdelta or timedlt>eddelta:
				continue
			if cmp(state, u'卖盘')==0:
				dataObj[j].sellvol += w5
				dataObj[j].sellct += 1
				#print "S:%d--%d %d" %(fltarray[j], dataObj[j].sellvol, dataObj[j].sellct)
			elif cmp(state, u'买盘')==0:
				dataObj[j].buyvol += w5
				dataObj[j].buyct += 1
				#print "B:%d--%d %d" %(fltarray[j], dataObj[j].buyvol, dataObj[j].buyct)

	#更新数据
	if "statistics" not in wb.get_sheet_names():
		print "No staticstics data"
		return
	ws = wb.get_sheet_by_name(name='statistics')
	max_row = ws.max_row

	fctime = ''
	qdate = g_sttime +"-"+ g_edtime
	ascid = 65
	row = max_row+2
	if cmp(fctime, '')==0:
		title = [qdate, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg', ]
	else:
		title = [qdate, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg', fctime]
	number = len(title)
	for i in range(0,number):
		cell = chr(ascid+i) + str(row)
		ws[cell] = title[i]

	dataObjLen = len(dataObj)
	for j in range(0, dataObjLen):
		list = []
		list.append(dataObj[j].volumn)
		
		buyvol = dataObj[j].buyvol
		buyct = dataObj[j].buyct
		
		sellvol = dataObj[j].sellvol
		sellct = dataObj[j].sellct
		
		list.append(buyvol)
		list.append(sellvol)
		list.append(buyct)
		list.append(sellct)
		if buyct==0:
			list.append(0)
		else:
			list.append(buyvol/buyct)
		if sellct==0:
			list.append(0)
		else:
			list.append(sellvol/sellct)
		list.append('')
		list.append(buyvol + sellvol)
		list.append(buyct + sellct)
		bsct = buyct + sellct
		if bsct==0:
			list.append(0)
		else:
			list.append((buyvol + sellvol)/bsct)
		
		row = row+1
		number = len(list)
		for i in range(0,number):
			cell = chr(ascid+i) + str(row)
			ws[cell] = list[i]
	wb.save(wkfile)
	
def updateFile(path, filename):
	fltlist = [200, 300, 600]
	wkfile = path +"/"+ filename

	dataObjLen = len(data_dic)
	if dataObjLen==0:
		print "没有数据"
		return
	
	wb = Workbook()
	ws = wb.active
	ws.title = 'statistics'
	
	ascid = 65
	row = 1
	for fltvol in fltlist:
		title = [fltvol, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg', ]
		number = len(title)
		for i in range(0,number):
			cell = chr(ascid+i) + str(row)
			ws[cell] = title[i]
		row += 1

		index = dataObjLen-1
		while index>=0:
			day_list = data_dic[index]
			index -= 1
			for data_list in day_list[1]:
				if data_list[0]!=fltvol:
					continue

				number = len(data_list)
				for i in range(0,number):
					cell = chr(ascid+i) + str(row)
					if i==0:
						ws[cell] = day_list[0]
					else:
						ws[cell] = data_list[i]
				row += 1
		#写完相同数据然后空一行
		row += 1
	wb.save(wkfile)

	
if __name__ == '__main__':
	pindex = len(sys.argv)
	if pindex<2:
		sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " [f=文件 | code date] HH-MM HH-MM\n")
		exit(1);

	bFile = 0
	idx = 1
	filename = ''
	param = sys.argv[idx]
	objParam = re.match(r'f=(.*)', param)
	if objParam:
		bFile = 1
		filename = objParam.group(1)
	else:
		ret, code = parseCode(sys.argv[idx])
		if ret!=0:
			exit(1)
		idx += 1
		today = datetime.date.today()
		ret,qdate = parseDate(sys.argv[idx], today)
		if ret!=0:
			exit(1)
		#print code, qdate
	idx += 1
	if pindex-idx<2:
		sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " [f=文件 | code date] HH-MM HH-MM\n")
		exit(1);

	sttime = sys.argv[idx]
	edtime = sys.argv[idx+1]
	if sttime=='.' and edtime=='.':
		sys.stderr.write("起始时间和结束时间不能都为'.'\n")
		exit(1);
	if sttime=='.':
		hour = 9
		minute = 15
		second = 0
		sttime = "09:15"
	else:
		objParam = re.match(r'(\d+):(\d+)', sttime)
		if objParam is None:
			sys.stderr.write("起始时间格式不正确\n")
			exit(1);
		hour = int(objParam.group(1))
		minute = int(objParam.group(2))
		second = 0
	stdelta = datetime.timedelta(hours=hour, minutes=minute, seconds=second).seconds
	g_sttime = sttime

	if edtime=='.':
		hour = 15
		minute = 01
		second = 0
		edtime = "15:00"
	else:
		objParam = re.match(r'(\d+):(\d+)', edtime)
		if objParam is None:
			sys.stderr.write("结束时间格式不正确\n")
			exit(1);
		hour = int(objParam.group(1))
		minute = int(objParam.group(2))
		second = 0
	eddelta = datetime.timedelta(hours=hour, minutes=minute, seconds=second).seconds
	g_edtime = edtime

	fileList = []
	if bFile==1:
		extname = filename.split('.')[-1]
		if cmp(extname,"xlsx")!=0:
			filename = filename + '.xlsx'
	else:
		filename = code +"_"+ qdate +".xlsx"
	path = prepath + code +"/"+ filename
	print path
	print os.path.exists(path) 

	if os.path.exists(path) is False:
		print "FIle not exist"
		exit(1)

	parseFile(path, dftArray, stdelta, eddelta)

'''

	if cmp('meg', '')!=0:
		stfile = '_'+ code +'__result.xlsx'
		updateFile(path, stfile)
'''		
