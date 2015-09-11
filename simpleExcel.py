# -*- coding:gbk -*-
import datetime
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

str = u'中文'
ws['A3'] = str.encode('utf8')

str = '汉子'
s1 = str.decode('gbk')
ws['A4'] = s1.encode('utf8') 


strline = u'成交时间,成交价,涨跌幅,价格变动,成交量,成交额,性质'
strObj = strline.split(',')
print strObj
ws.append(strObj)

# Save the file
wb.save("z_excel.xlsx")
