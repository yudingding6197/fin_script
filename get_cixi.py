#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import time
import datetime
import tushare as ts
import pandas as pd
import shutil
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from internal.common import *

#������ִ��debug\newstk_get.py,�õ����յ�new

#��ȡ����еļ۸�
def get_xg_fx():
	sheet_st = "Sheet"
	wkfile = "../data/entry/xingu/faxing.xlsx"
	wkfile1 = "../data/entry/xingu/faxing_re.xlsx"

	wb = load_workbook(wkfile)
	ws = wb.get_sheet_by_name(sheet_st)	
	first_list = []
	item_list = []
	xg_df = pd.DataFrame()
	for rx in range(1,ws.max_row+1):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value

		if rx==1:
			first_list = [w1,w2,w3,w4,w5,w6]
			continue

		temp_list = [w1,w2,w3,w4,w5,w6]
		item_list.append(temp_list)
	df1 = pd.DataFrame(item_list, columns=first_list)
	xg_df = xg_df.append(df1)

	item_list = []
	wb = load_workbook(wkfile1)
	ws = wb.get_sheet_by_name(sheet_st)	
	for rx in range(2,ws.max_row+1):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value

		temp_list = [w1,w2,w3,w4,w5,w6]
		item_list.append(temp_list)
	if len(item_list)>0:
		df1 = pd.DataFrame(item_list, columns=first_list)
		xg_df = xg_df.append(df1)

	xg_df = xg_df.set_index('code')
	#print xg_df.ix['001979'][3]

	return xg_df

#�õ���ǰ�����������
def get_hist_cx():
	sheet_st = "Sheet"
	wkfile = "../data/entry/cixin/cx_anly_latest.xlsx"
	if os.path.isfile(wkfile) is False:
		return None

	wb = load_workbook(wkfile)
	ws = wb.get_sheet_by_name(sheet_st)	
	first_list = []
	item_list = []
	#hist_df = pd.DataFrame()
	for rx in range(1,ws.max_row+1):
		temp_list = []
		for j in range(1, 19):
			wobj = ws.cell(row = rx, column = j).value
			temp_list.append(wobj)
		if rx==1:
			first_list = temp_list
			first_list[0] = 'code'
			first_list[6] = 'ipo_date'
			continue
		item_list.append(temp_list)
	#df1 = pd.DataFrame(item_list, columns=first_list)
	if len(item_list)>0:
		#df1 = pd.DataFrame(item_list, columns=first_list)
		#hist_df = hist_df.append(df1)
		return item_list
	else:
		return None

	#hist_df = hist_df.set_index('code')
	return hist_df

#�õ������ϵ�Item
def get_xg_trade(xg_trade_list, xg_df):
	sheet_st = "Sheet"
	wkfile = "../data/entry/trade/ns_info.xlsx"
	if os.path.isfile(wkfile) is False:
		print "Error: No File:", wkfile
		return None

	wb = load_workbook(wkfile)
	ws = wb.get_sheet_by_name(sheet_st)	
	item_list = []
	first_list = ['code','name','timeToMarket','outstanding','totals']
	new_st_dt = pd.DataFrame()
	for rx in range(1,ws.max_row+1):
		if rx==1:
			continue
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		ipo_date = ws.cell(row = rx, column = 3).value
		liutong_gb = ws.cell(row = rx, column = 4).value
		zong_gb = ws.cell(row = rx, column = 5).value

		#print xg_df.ix[w1][u'������']
		#ipo_date = xg_df.ix[w1]['timeToMarket']
		#liutong_gb = xg_df.ix[w1]['outstanding']
		#zong_gb = xg_df.ix[w1]['totals']
		'''
		'''
		temp_list = [w1,w2,ipo_date,liutong_gb,zong_gb]
		item_list.append(temp_list)
		xg_trade_list.append(w1)
	df1 = pd.DataFrame(item_list, columns=first_list)
	new_st_dt = new_st_dt.append(df1)
	new_st_dt = new_st_dt.set_index('code')
	return new_st_dt

def append_stk_info(stockInfo, ws):
	global excel_row

	k = 0
	ascid = 65
	number = len(stockInfo)
	for k in range(0,number):
		cell = chr(ascid+k) + str(excel_row)
		ws[cell] = stockInfo[k]
	excel_row += 1

def get_stock_data(code):
	ret, pcode = parseCode(code)
	if ret==-1:
		print code, "Parse it fail"
		return (-1, "", "")

	start=time.clock()
	LOOP_COUNT = 0
	trdf = None
	while LOOP_COUNT<3:
		try:
			trdf = ts.get_realtime_quotes(code)
		except:
			LOOP_COUNT += 1
			time.sleep(0.5)
			end = time.clock()
			if LOOP_COUNT==3:
				print code,"Exception=",LOOP_COUNT," ",end-start
		else:
			break
	if trdf is not None:
		volstr = trdf.iloc[0,10]
		rt_name = trdf.iloc[0,0]
		return (0, volstr, rt_name)

	urllink = "http://push2.gtimg.cn/q=" + pcode
	LOOP_COUNT = 0
	stockData = None
	while LOOP_COUNT<3:
		try:
			req = urllib2.Request(urllink)
			stockData = urllib2.urlopen(req, timeout=5).read()
		except:
			LOOP_COUNT += 1
			end = time.clock()
			if LOOP_COUNT==3:
				print code,"push2 gtimg Exception=",LOOP_COUNT," ",end-start
		else:
			break;
	if stockData is None:
		return (-1, "", "")
	stockObj = stockData.split('~')
	if len(stockObj)<2:
		return (-1, "", "")

	v1 = stockObj[1].decode('gbk')
	v2 = stockObj[6]
	#print code, v1, v2
	return (0, v2, v1)

def parse_item_data(type, code, row, xg_df, ws, hist_df):
	stockInfo = []
	if type==1:
		name = row[0]
		ipo_date = long(row['timeToMarket'])
		ltgb = float(row['outstanding'])/10000
		liutong_gb = round(ltgb, 2)
		zgb = float(row['totals'])/10000
		zong_gb = round(zgb, 2)
	elif type==2:
		name = row[0].decode('utf8')
		ipo_date = row['timeToMarket']
		liutong_gb = row['outstanding']
		zong_gb = row['totals']
	else:
		return
	#print type(ipo_date), type(liutong_gb), type(zong_gb)
	#print type(ipo_date) ��Ȼ�� long ����
	#print code, ipo_date, liutong_gb, zong_gb
	if long(ipo_date)==0:
		return
	trade_string = str(long(ipo_date))
	trade_date = datetime.datetime.strptime(trade_string, '%Y%m%d').date()
	delta = trade_date - base_date
	if delta.days<0:
		return

	#������������2�����
	if type==2 and hist_df is not None:
		#today = datetime.date.today()
		#cmp_delta = today - trade_date
		tmpInfo = []
		for j in range(0, len(hist_df)):
			if hist_df[j][0]!=code:
				continue
			tmpInfo = hist_df[j]
			break
		if len(tmpInfo)==0:
			#������������û�б��棬��������
			print code, "latest data not include the stk"
		elif tmpInfo[3]==0:
			pass
		else:
			append_stk_info(tmpInfo, ws)
			return

	#���ÿֻ����ÿ�콻������
	LOOP_COUNT = 0
	tddf = None
	while LOOP_COUNT<3:
		try:
			tddf = ts.get_k_data(code, autype='bfq')
		except:
			LOOP_COUNT += 1
			time.sleep(0.5)
		else:
			break;
	if tddf is None:
		print "Timeout to get k data of " + code +", Quit"
		exit(0)

	b_open = 0
	b_break = 0
	b_log_cxkb = 0
	yzzt_day = 0
	last_close = 0.0
	td_total = len(tddf)
	fengban_vol = 0		#�������
	fengliu_prop = 0.0
	last_day_vol = 0.0
	turnover = 0.0
	cxkb_date = 0
	kbzt_days = 0		#�����������ͣ�����Ƿ�����ͣ
	kbczt_days = 0
	kbdt_days = 0		#�����������ͣ�����Ƿ�����ͣ
	for tdidx,tdrow in tddf.iterrows():
		open = tdrow[1]
		close = tdrow[2]
		high = tdrow['high']
		low = tdrow['low']
		last_day_vol = tdrow['volume']
		#high == low ��ζYZZT
		#print code,name,kbczt_days,kbzt_days,tdrow['date'],high,low
		if high==low:
			yzzt_day += 1
			last_close = close
			if b_open==1:
				kbczt_days += 1
				kbzt_days += 1
			continue

		#�¹ɵ�һ����� high!=low
		if yzzt_day!=0:
			b_open = 1
			opn_date_str = tdrow['date']
			#������ͣ����
			zt_price = last_close * 1.0992
			dt_price = last_close * 0.9005
			#print code,name,last_close,close,zt_price,high,kbczt_days,kbzt_days
			if (close>=zt_price or high>=zt_price) and (kbdt_days==0):
				#ͨ�����㿪������ͣ����ͣ�����������ж�
				#�����һ�����ͣ���ڶ�����ͣ������Ϊ������ͣ
				if kbczt_days==kbzt_days and close>=zt_price:
					kbzt_days += 1
				kbczt_days += 1
			elif close<=dt_price and kbczt_days==0 and kbzt_days==0:
				kbdt_days += 1
			else:
				b_break = 1
			b_log_cxkb = 1
		else:
			#��������¹ɣ������߿ڡ����Ϲɷݵ�
			if open<=close and close==high and low==open:
				yzzt_day += 1
			else:
				b_open = 1
				b_break = 1
				b_log_cxkb = 1
				opn_date_str = tdrow['date']
		last_close = close
		
		#���ܿ��壬���ǿ��ܻ����������ZT or DT��������Ҫ��¼��������
		if b_log_cxkb==1:
			if cxkb_date==0:
				opn_date_int = ''.join(opn_date_str.split('-'))
				cxkb_date = int(opn_date_int)
		if b_break==1:
			break

	#ֻ��û�д򿪵�CX���ż����⼸�����ݣ��򿪵ľͺ���
	if b_open==0:
		LOOP_COUNT = 0
		trdf = None

		ret, volstr, rt_name = get_stock_data(code)
		if ret==-1:
			print code, "Timeout to get item data"
			exit(0)

		#volstr = trdf.iloc[0,10]
		if volstr.isdigit() is True:
			fengban_vol = int(volstr)
			fengliu_prop = fengban_vol/(liutong_gb*10000)
			turnover = last_day_vol/(liutong_gb*10000)
		#�ж����֣�����CX���� N...��ʽ
		#rt_name = trdf.iloc[0,0]
		if name!=rt_name:
			name = rt_name

	#׷������,��ͨ��ֵ������ֵ
	liutong_sz = liutong_gb*last_close
	zong_sz = zong_gb*last_close
	stockInfo.append(code)
	stockInfo.append(name)
	if xg_df is None:
		stockInfo.append(None)
	else:
		if code in xg_list:
			stockInfo.append(xg_df.ix[code][3])
		else:
			print code,name, "No in list"
	stockInfo.append(b_open)
	stockInfo.append(yzzt_day)
	stockInfo.append(last_close)
	stockInfo.append(ipo_date)
	stockInfo.append(cxkb_date)
	stockInfo.append(kbzt_days)
	stockInfo.append(kbczt_days)
	stockInfo.append(liutong_gb)
	stockInfo.append(round(liutong_sz,2))
	stockInfo.append(zong_gb)
	stockInfo.append(round(zong_sz,2))
	stockInfo.append(fengban_vol)
	stockInfo.append(round(fengliu_prop,2))
	stockInfo.append(round(turnover,2))
	stockInfo.append(kbdt_days)
	#print stockInfo
	append_stk_info(stockInfo, ws)

# Main
if __name__ =='__main__':
	cmp_string = "20150201"
	base_date = datetime.datetime.strptime(cmp_string, '%Y%m%d').date()

	prepath = "../data/"
	prepath1 = "../data/entry/cixin/"
	LOOP_COUNT = 0
	df = None

	xg_df = get_xg_fx()
	xg_list = list(xg_df.index)

	new_st_list = []
	new_st_dt = get_xg_trade(new_st_list, xg_df)

	#�����µ��¹ɽ����л������
	hist_df = get_hist_cx()
	#print hist_df.head(15)

	#�õ�basic info
	while LOOP_COUNT<3:
		try:
			df = ts.get_stock_basics()
		except:
			LOOP_COUNT += 1
			time.sleep(0.5)
		else:
			break;
	if df is None:
		print "Timeout to get stock basic info"
		exit(0)
	df1 = df[df.timeToMarket>0]
	df1 = df1.sort_values(['timeToMarket'], 0, False)
	st_index = df1.index
	st_bas_list=list(st_index)
	#df1 = df1[:10]

	st_list = []
	for i in range(0, len(new_st_list)):
		#print "======="
		if new_st_list[i] in st_bas_list:
			pass
		else:
			print new_st_list[i], " Not Exist"
			st_list.append(new_st_list[i])
	#print st_list

	index = -1
	wb = Workbook()
	# grab the active worksheet
	ws = wb.active
	strline = u'����,����,���м�,����,�����,����,������,������,����ZT,����CZT,��ͨ�ɱ�,��ͨ��ֵ,�ܹɱ�,����ֵ,�ⵥ����,�ⵥ��ͨ��,������,����DT'
	strObj = strline.split(u',')
	ws.append(strObj)
	#�����������иı�
	ws.auto_filter.ref = "A1:R1"
	excel_row = 2

	#����н��շ��и���
	for st_item in st_list:
		row = new_st_dt.ix[st_item]
		parse_item_data(1, st_item, row, xg_df, ws, hist_df)

	for code,row in df1.iterrows():
		parse_item_data(2, code, row, xg_df, ws, hist_df)
	#print "Final excel_row=",excel_row
	filexlsx = prepath + "cixin_analyze.xlsx"
	wb.save(filexlsx)

	today = datetime.date.today()
	cur=datetime.datetime.now()
	qdate = '%04d-%02d-%02d' %(today.year, today.month, today.day)
	filexlsx1 = prepath1 + "cx_anly_"+ qdate
	filexlsx1 = '%s#%02d-%02d.xlsx' %(filexlsx1, cur.hour, cur.minute)
	wb.save(filexlsx1)
	cpfile = prepath1 + "cx_anly_latest.xlsx"
	shutil.copy(filexlsx1, cpfile)
	print "Get CX Finished"
