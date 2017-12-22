#!/usr/bin/env python
# -*- coding:gbk -*-
import sys
import re
import os
import time
import string
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from internal.common import *

#��ָ�����ڵ� statistics ��B/S��ͳ�����ݣ��ŵ���Ӧ����_parserĿ¼��
#��000520���ŵ� ../data/sz000520_parser/�µ��ļ���
prepath = '../data/'
sheetName = "statistics"
allFileNum = 0
upd_flag = 0
volumeList = []
volumeList1 = []

def printPath(level, path,fileList):
	global allFileNum
	# �����ļ��У���һ���ֶ��Ǵ�Ŀ¼�ļ���
	dirList = []
	# �����ļ�
	#fileList = []
	# ����һ���б����а�����Ŀ¼��Ŀ������(google����)
	files = os.listdir(path)
	# �����Ŀ¼����
	dirList.append(str(level))
	for f in files:
		if(os.path.isdir(path + '/' + f)):
			# �ų������ļ��С���Ϊ�����ļ��й���
			if(f[0] == '.'):
				pass
			else:
				# ��ӷ������ļ���
				dirList.append(f)
		if(os.path.isfile(path + '/' + f)):
			# ����ļ�
			fileList.append(f)
	# ��һ����־ʹ�ã��ļ����б��һ�����𲻴�ӡ
	''' Ŀǰ����Ҫ���ļ���
	i_dl = 0
	for dl in dirList:
		if(i_dl == 0):
			i_dl = i_dl + 1
		else:
			# ��ӡ������̨�����ǵ�һ����Ŀ¼
			print '-' * (int(dirList[0])), dl
			# ��ӡĿ¼�µ������ļ��к��ļ���Ŀ¼����+1
			printPath((int(dirList[0]) + 1), path + '/' + dl)
	for fl in fileList:
		# ��ӡ�ļ�
		#print '-' * (int(dirList[0])), fl
		# ������һ���ж��ٸ��ļ�
		allFileNum = allFileNum + 1
		file = os.open(fl, "r")
	'''
	
	'''
	printPath(1, path, fileList)
	for fl in fileList:
		wkfile = path +"//"+ fl
		wb = load_workbook(wkfile)
		print "Worksheet range(s):", wb.get_named_ranges()
		print "Worksheet name(s):", wb.get_sheet_names()
		if sheetName not in wb.get_sheet_names():
			print "NONNN"
			continue

		sheet_ranges_r = wb.get_sheet_by_name(name=sheetName)
		print "YYYY:"
		print sheet_ranges_r
	'''	

# �����洢���ݵ��ֵ� 
data_dic = []
bCheckDate = 0

def parseFile(path, filename):
	global upd_flag
	wkfile = path +"/"+ filename

	wb = load_workbook(wkfile)
	#print "Worksheet range(s):", wb.get_named_ranges()
	#print "Worksheet name(s):", wb.get_sheet_names()
	if sheetName not in wb.get_sheet_names():
		return

	ws = wb.get_sheet_by_name(name=sheetName)
	max_column = ws.max_column
	if max_column<7:
		print "column(%d) too short, please check" % max_column
		return

	#�����ݴ浽�ֵ���
	pid = ''
	data_list = []
	for rx in range(1,ws.max_row):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value
		w7 = ws.cell(row = rx, column = 7).value
		w8 = ws.cell(row = rx, column = 8).value
		w9 = ws.cell(row = rx, column = 9).value

		if (w1 is None) and (w2 is None) and (w3 is None) and (w4 is None)\
			and (w5 is None) and (w6 is None) and (w7 is None):
			break
		if (w1 is None) or (w2 is None) or (w3 is None) or (w4 is None)\
			or (w5 is None) or (w6 is None) or (w7 is None):
			print rx,"ĳ���¼ΪNone������ȷ", w1, w2, w3, w4, w5, w6, w7 
			continue
		temp_list = [w1,w2,w3,w4,w5,w6,w7,w8,w9]
		if rx==1:
			pid = w1
			continue
		#print temp_list
		data_list.append(temp_list)
		if upd_flag==0:
			if w1!=0:
				volumeList.append(w1)
			volumeList1.append(w1)
		else:
			if volumeList1[rx-2]!=w1:
				print "Warning: Line(%d) value not match (%d,%d)" %(rx, volumeList1[rx-1],w1)
	if upd_flag==0:
		upd_flag=1
	#print volumeList, volumeList1

	#�õ�������Ϣ�����̼ۣ����̼۵�
	ws = wb.get_sheet_by_name(name='Sheet')
	max_column = ws.max_row
	w1 = ws.cell(row = 2, column = 8).value
	w2 = ws.cell(row = 2, column = 9).value
	w3 = ws.cell(row = 2, column = 10).value
	w4 = ws.cell(row = 2, column = 11).value
	#�������̳ɽ���
	w5 = ws.cell(row = max_column, column = 5).value
	w6 = ws.cell(row = 2, column = 5).value
	#�ɽ���
	w7 = ws.cell(row = 2, column = 14).value
	w7 = int(w7)
	'''
	if w2<0:
		f1 = '%02.02f'%(w2)
	else:
		f1 = '%02.02f'%(w2)
	v3 = '%02.02f'%(w3)
	v4 = '%02.02f'%(w4)
	v1 = '%02.02f'%(w1)
	'''
	#�ֱ��ǣ����ռ�,���̼�,���̼�,�Ƿ�,���̳ɽ�,���̳ɽ�,�ɽ���
	#trade_info = [v3, v4, v1, f1]
	trade_info = [w3, w4, w1, w2, w5, w6, w7]

	max_column = ws.max_column
	if cmp(pid, '')!=0:
		day_list = [pid, data_list, trade_info]
		data_dic.append(day_list)

def createResultFile(path, filename):
	wkfile = path +"/"+ filename

	dataObjLen = len(data_dic)
	if dataObjLen==0:
		print "û������"
		return
	
	wb = Workbook()
	ws = wb.active
	ws.title = sheetName
	addStatVolume(ws, 0)

	ws = wb.create_sheet()
	ws.title = 'detail'
	addStatVolume(ws, 1)
	wb.save(wkfile)
	
def addStatVolume(ws, flag):
	dataObjLen = len(data_dic)
	row = 1
	ascid = 65
	title = []

	i = 0
	bsIndex = 0
	strline = u'ǰ�ռ�,���̼�,���̼�,�ǵ���,���̳ɽ�,���̳ɽ�,�ɽ���'
	strObj = strline.split(u',')
	strObjLen = len(strObj)
	title.append('Date')
	if flag==0:
		title.append('FLUC')
	#��detailҳ�����ڼ���
	elif flag==1:
		for j in range(0,strObjLen):
			title.append(strObj[j])
	#�õ�B/S�е���ʼindex
	bsIndex = len(title)

	#���B200,S200��title
	for fltvol in volumeList1:
		bs = 'B%d'%(fltvol)
		title.append(bs)
		bs = 'S%d'%(fltvol)
		title.append(bs)
		if fltvol==200 or fltvol==300 or (flag==1 and fltvol!=0):
			percent = 'P%d'%(fltvol)
			title.append(percent)

	#��ͳ��ҳ�����ڼ���
	if flag==0:
		for j in range(0,strObjLen):
			title.append(strObj[j])

	number = len(title)
	for i in range(0,number):
		cell = chr(ascid+i) + str(row)
		ws[cell] = title[i]

	#���ַ�ʽ���� list ����
	dlen = len(volumeList1)
	buy  = [0 for x in range(dlen)]
	sell = [0 for x in range(dlen)]
	stat = [0] * dlen
	index = dataObjLen-1
	row += 1
	while index>=0:
		i = 0
		day_list = data_dic[index]
		index -= 1

		cell = chr(ascid+i) + str(row)
		ws[cell] = day_list[0]
		i += 1

		if flag==0:
			#��ӵ��콻���ǵ���
			trade_info = day_list[2]
			cell = chr(ascid+i) + str(row)
			ws[cell] = trade_info[3]
			i += 1
		elif flag==1:
			#��ӵ��콻����Ϣ
			trade_info = day_list[2]
			for j in range(0, len(trade_info)):
				cell = chr(ascid+i) + str(row)
				ws[cell] = trade_info[j]
				i += 1

		j = 0
		dayBVolume = 0
		daySVolume = 0
		for data_list in day_list[1]:
			buy[j] = buy[j] + data_list[1]
			cell = chr(ascid+i) + str(row)
			ws[cell] = data_list[1]
			i += 1

			sell[j] = sell[j] + data_list[3]
			cell = chr(ascid+i) + str(row)
			ws[cell] = data_list[3]
			i += 1

			#�õ�������������
			if data_list[0]==0:
				dayBVolume = data_list[1]
				daySVolume = data_list[3]
				#print dayBVolume, daySVolume

			#��Ӱٷֱ�
			if data_list[0]==200 or data_list[0]==300 or (flag==1 and data_list[0]!=0):
				stat[j] = 1
				if dayBVolume==0:
					buyPerc = 0
				else:
					buyPerc = round(float(data_list[1]) * 100 / (dayBVolume), 2)
				if daySVolume==0:
					sellPerc = 0
				else:
					sellPerc = round(float(data_list[3]) * 100 / (daySVolume), 2)
				value = '%2.1f-%2.1f'%(buyPerc, sellPerc)
				cell = chr(ascid+i) + str(row)
				ws[cell] = value
				i += 1
			j += 1

		if flag==0:
			#��ӵ��콻����Ϣ
			trade_info = day_list[2]
			for j in range(0, len(trade_info)):
				cell = chr(ascid+i) + str(row)
				ws[cell] = trade_info[j]
				i += 1

		row += 1

	index = 0
	cell = chr(ascid+index) + str(row)
	ws[cell] = "Total"
	index += bsIndex
	number = len(buy)
	#����ͳ������ɽ��������������ɽ���������
	#��Щ�л���������������(����)/�ܹ�����(����)��ռ��
	for i in range(0, number):
		cell = chr(ascid+index) + str(row)
		ws[cell] = buy[i]
		index += 1
		cell = chr(ascid+index) + str(row)
		ws[cell] = sell[i]
		index += 1
		if i==0:
			dayBVolume = buy[i]
			daySVolume = sell[i]
			continue
		if stat[i]==1:
			if dayBVolume==0:
				buyPerc = 0
			else:
				buyPerc = round(float(buy[i]) * 100 / (dayBVolume), 2)
			if daySVolume==0:
				sellPerc = 0
			else:
				sellPerc = round(float(sell[i]) * 100 / (daySVolume), 2)
			value = '%2.1f-%2.1f'%(buyPerc, sellPerc)
			cell = chr(ascid+index) + str(row)
			ws[cell] = value
			index += 1
	row += 1

	#��������(����)/�ܹ�����(����) ��ռ��
	totalBVol = buy[0]
	totalSVol = sell[0]
	column = bsIndex+2
	for i in range(0, number):
		if i==0:
			continue;

		if totalBVol==0:
			buyPerc = 0
		else:
			buyPerc = round(float(buy[i]) * 100 / totalBVol, 2)
		cell = chr(ascid+column) + str(row)
		ws[cell] = buyPerc
		column += 1

		if totalSVol==0:
			sellPerc = 0
		else:
			sellPerc = round(float(sell[i]) * 100 / totalSVol, 2)
		cell = chr(ascid+column) + str(row)
		ws[cell] = sellPerc
		column += 1

		if stat[i]==1:
			column += 1
	row += 1

	#��������(����)/������������ ��ռ��
	column = bsIndex
	totoalVol = totalBVol+totalSVol
	for i in range(0, number):
		if totoalVol==0:
			buyPerc = 0
		else:
			if i==0:
				buyPerc = round(float(totalBVol) * 100 / totoalVol, 2)
			else:
				buyPerc = round(float(buy[i]) * 100 / totoalVol, 2)
		cell = chr(ascid+column) + str(row)
		ws[cell] = buyPerc
		column += 1

		if totoalVol==0:
			sellPerc = 0
		else:
			if i==0:
				sellPerc = round(float(totalSVol) * 100 / totoalVol, 2)
			else:
				sellPerc = round(float(sell[i]) * 100 / totoalVol, 2)
		cell = chr(ascid+column) + str(row)
		ws[cell] = sellPerc
		column += 1

		if stat[i]==1:
			column += 1
	row += 1

	if flag==0:
		ws.auto_filter.ref = "A1:B1"
	elif flag==1:
		ws.auto_filter.ref = "A1:H1"

def statByVolume(path, filename):
	wkfile = path +"/"+ filename

	dataObjLen = len(data_dic)
	if dataObjLen==0:
		print "û������"
		return
	
	wb = Workbook()
	ws = wb.active
	ws.title = sheetName
	
	ascid = 65
	row = 1
	for fltvol in volumeList:
		title = [fltvol, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg']
		number = len(title)
		for i in range(0,number):
			cell = chr(ascid+i) + str(row)
			ws[cell] = title[i]
		row += 1

		index = dataObjLen-1
		while index>=0:
			day_list = data_dic[index]
			index -= 1
			for data_list in day_list[1]:
				if data_list[0]!=fltvol:
					continue

				number = len(data_list)
				for i in range(0,number):
					cell = chr(ascid+i) + str(row)
					if i==0:
						ws[cell] = day_list[0]
					else:
						ws[cell] = data_list[i]
				row += 1
		#д����ͬ����Ȼ���һ��
		row += 1
	wb.save(wkfile)

	
if __name__ == '__main__':
	pindex = len(sys.argv)
	if pindex==2 or pindex==4:
		pass
	else:
		sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " ���� [��ʼʱ��<YYYYMMDD or MMDD>] [����ʱ��<YYYYMMDD or MMDD>]\n")
		exit(1);

	code = sys.argv[1]
	if (len(code) != 6):
		sys.stderr.write("Len should be 6\n")
		exit(1);

	ret, code=parseCode(code)
	if ret!=0:
		exit(1);

	bCheckDate = 0
	styear = 0
	stmonth = 0
	stday = 0
	edyear = 0
	edmonth = 0
	edday = 0
	today = datetime.date.today()
	if pindex==4:
		ret,stdate= parseDate(sys.argv[2], today)
		if ret==-1:
			exit(1)
		startObj = time.strptime(stdate, "%Y-%m-%d")

		edstr = sys.argv[3]
		if cmp(edstr, '.')==0:
			eddate = '%04d-%02d-%02d' %(today.year, today.month, today.day)
		else:
			ret,eddate = parseDate(edstr, today)
			if ret==-1:
				exit(1)
		endObj = time.strptime(eddate, "%Y-%m-%d")
		#���ü���־λ
		bCheckDate = 1

	fileList = []
	path = prepath + code
	if not os.path.isdir(path):
		print "'"+ path +"' not a dir"
		exit(1)

	resPath = prepath + code + "_parse"
	if not os.path.isdir(resPath):
		os.makedirs(resPath)

	for (dirpath, dirnames, filenames) in os.walk(path):  
		#print('dirpath = ' + dirpath)
		i = 0
		for filename in filenames:
			extname = filename.split('.')[-1]
			if cmp(extname,"xlsx")!=0:
				continue
			prename = code+"_"
			if cmp(filename[0:9], prename[0:9])!=0:
				continue
			if bCheckDate==1:
				fileDate = filename[9:(len(filename)-5)]
				if len(fileDate)!=10:
					continue

				try:
					curObj = time.strptime(fileDate, "%Y-%m-%d")
				except:
					print "Invalid date:",fileDate
					continue

				if curObj<startObj or curObj>endObj:
					continue
			#print path, filename
			parseFile(path, filename)
			i += 1
		#�����õ����ļ��е��ļ����������ļ������ļ�
		break

	fctime = '_%02d%02d_%02d%02d'%(startObj.tm_mon, startObj.tm_mday, endObj.tm_mon, endObj.tm_mday)
	stfile = 'z_'+ code + fctime + '.xlsx'
	createResultFile(resPath, stfile)
	#stfile = 'z_'+ code +'_result_'+fctime+'.xlsx'
	#statByVolume(path, stfile)
