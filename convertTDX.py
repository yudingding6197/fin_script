# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib
import urllib2
import binascii
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
import internal.common

#将TDX的数据转换为excel

addcsv = 0
prepath = "..\\Data\\TDX\\"
convAll = 0
pindex = len(sys.argv)
if pindex<2:
	sys.stderr.write("转换通达信的数据到Excel中, 如果没有指定日期，转换所有文件\n")
	sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " 代码 [日期]\n")
	sys.stderr.write("日期:d=YYYY-MM-DD | d=MM-DD | d=MM-DD~MM-DD | d=YYYY-MM-DD~YYYY-MM-DD] \n")
	sys.stderr.write("     前两种是单独日期, 后两种是范围\n")
	exit(1);

code = sys.argv[1]
if (len(code) != 6):
	sys.stderr.write("Len should be 6\n")
	exit(1);
	
if pindex==2:
	convAll = 1

qdate = ""
convType = 0
if (pindex==3):
	today = datetime.date.today()
	defdate = sys.argv[2]
	dateObj = defdate.split('~')
	dtLen = len(dateObj)
	print "dlen=%d" %(dtLen)
	if (len(dateObj)==1):
		convType = 0
		qdate = dateObj[0]
		dateObj = re.match(r'^(\d{4})-(\d+)-(\d+)', qdate)
		if (dateObj is None):
			dateObj = re.match(r'^(\d+)-(\d+)', qdate)
			if (dateObj is None):
				print "非法日期格式：" +qdate+ ",期望格式:YYYY-MM-DD or MM-DD"
				exit(1);
			else:
				year = today.year
				month = int(dateObj.group(1))
				day = int(dateObj.group(2))
		else:
			year = int(dateObj.group(1))
			month = int(dateObj.group(2))
			day = int(dateObj.group(3))
		qdate = '%04d%02d%02d' %(year, month, day)
		print qdate
	else:
		convType = 1
		sdate = dateObj[0]
		edate = dateObj[1]
		dateObj = re.match(r'^(\d{4})-(\d+)-(\d+)', sdate)
		if (dateObj is None):
			dateObj = re.match(r'^(\d+)-(\d+)', sdate)
			if (dateObj is None):
				print "非法日期格式：" +sdate+ ",期望格式:YYYY-MM-DD or MM-DD"
				exit(1);
			else:
				year = today.year
				month = int(dateObj.group(1))
				day = int(dateObj.group(2))
		else:
			year = int(dateObj.group(1))
			month = int(dateObj.group(2))
			day = int(dateObj.group(3))
		sdate = '%04d%02d%02d' %(year, month, day)
		
		dateObj = re.match(r'^(\d{4})-(\d+)-(\d+)', edate)
		if (dateObj is None):
			dateObj = re.match(r'^(\d+)-(\d+)', edate)
			if (dateObj is None):
				print "非法日期格式：" +edate+ ",期望格式:YYYY-MM-DD or MM-DD"
				exit(1);
			else:
				year = today.year
				month = int(dateObj.group(1))
				day = int(dateObj.group(2))
		else:
			year = int(dateObj.group(1))
			month = int(dateObj.group(2))
			day = int(dateObj.group(3))
		edate = '%04d%02d%02d' %(year, month, day)
		
		print sdate
		print edate


path = prepath + code
print path
dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(--|\+?\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D')
if convAll==0:
	if convType==0:
		filename = qdate + "-" + code + ".txt"
		filepath = path + "\\" + filename
		countl = 0
		with open(filepath, 'r') as recFile:
			for line in recFile:
				countl += 1
				if countl>20:
					break

				recObj = re.match(r'^(\d+\:\d+)[	 ]+(\d+\.\d+)[	 ]+(\d+)[	 ]+(\d+)[	 ]+(\S*)', line)
				if recObj:
					#<tr ><th>11:29:36</th><td>14.56</td><td>-3.13%</td><td>+0.01</td><td>9</td><td>13,104</td><th><h6>卖盘</h6></th></tr>
					print (recObj.groups())
					newline = " " + recObj.group(1) + ":00<>" + recObj.group(2) + "</td><td>" + "+0%<>" + "+0.02<>321</td><td>" + recObj.group(3) + "</td><th><h1>" + "买盘" + "</h1>"
					print newline
					#key = re.match(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(卖盘|买盘|中性盘)\D', newline)
					key = dtlRe.match(line)
					if (key is None):
						print "None"
						continue
					else:
						print key.groups()
					
					'''
					curtime = key.group(1)
					curvol = int(key.group(5))
					#记住当前页第一个的时间
					if (bFtime==0):
						timeobj = re.search(curtime, pageFtime)
						if timeobj:
							break
						pageFtime = curtime
						bFtime = 1

					timeobj = re.search(curtime, lasttime)
					if (timeobj and curvol==lastvol):
						pass
					else:
						curprice = key.group(2)
						fluctuate = key.group(4)
						lasttime = curtime
						lastvol = curvol
						amount = key.group(6)
						obj = amount.split(',')
						amount = ''.join(obj)

						intamount = int(key.group(5))
						updatestate = key.group(7)
						state = key.group(7)
						if cmp(state, '卖盘')==0:
							handle_volumn(intamount, dataObj, 2)
						elif cmp(state, '买盘')==0:
							handle_volumn(intamount, dataObj, 1)
						elif cmp(state, '中性盘')==0:
							ret = handle_middle_volumn(intamount, dataObj, curtime, fluctuate, key.group(3))
							if ret==1:
								state = '买盘'
							elif ret==2:
								state = '卖盘'

						if addcsv==1:
							strline = curtime +","+ key.group(2) +","+ key.group(3) +","+ key.group(4) +","+ key.group(5) +","+ amount +","+ key.group(7) + "\n"
							fcsv.write(strline)

						totalline += 1
						row = totalline+1
						cell = 'A' + str(row)
						ws[cell] = curtime
						cell = 'B' + str(row)
						ws[cell] = float(key.group(2))
						cell = 'C' + str(row)
						ws[cell] = key.group(3)
						cell = 'D' + str(row)
						if (fluctuate=='--'):
							ws[cell] = key.group(4)
						else:
							ftfluct = float(fluctuate)
							ws[cell] = ftfluct
						cell = 'E' + str(row)
						ws[cell] = int(key.group(5))
						cell = 'F' + str(row)
						ws[cell] = int(amount)
						cell = 'G' + str(row)
						s1 = state.decode('gbk')
						ws[cell] = s1
						
						if (row==2 and bhist==1):
							ascid = 72
							number = len(stockInfo)
							for j in range(0,number):
								cell = chr(ascid+j) + str(row)
								ws[cell] = stockInfo[j]
					'''	
				#s = binascii.b2a_hex(line)
				#print s
				#print "---%d:'%s'" %(countl, line)
		recFile.close()

	elif convType==1:
		print "还未实现。。。"
		pass
elif convAll==1:
	for (dirpath, dirnames, filenames) in os.walk(path):  
		print('dirpath = ' + dirpath)
		i = 0
		for filename in filenames:
			extname = filename.split('.')[-1]
			if cmp(extname,"txt")!=0:
				continue

			print filename
			#parseFile(path, filename)
			i += 1
			
		#仅仅得到父文件夹的文件，忽略子文件夹下文件
		break;
	print "还未实现。。。"


#internal.common.handle_data(addcsv, prepath, 0, url, code, qdate, sarr)


