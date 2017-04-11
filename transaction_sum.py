# -*- coding:gbk -*-
import sys
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

# Main cc:持仓， last_cc:每天最后一笔的成交总计
path = "..\\buy_sell\\"
file_list = []
filter_item = "A1:K1"
c_list = ['date','time','code','name','op','vol','price','amount','cc','last_cc']
cc_idx = c_list.index('last_cc')
for f in os.listdir(path):
	if os.path.isfile(path + f) is False:
		continue
	obj = re.match(r'table_(\d+).xlsx', f)
	if obj is None:
		print "Not matched file:",f
		continue
	file_list.append(f)

df = pd.DataFrame()
st_date = file_list[0][6:12]
#print file_list[-1]
ed_date = file_list[-1][6:12]
#print st_date,ed_date
for file in file_list:
	dt_str = file[6:12]
	if dt_str.isdigit() is False:
		print "Invalid file(%s) or date(%s)" % (file, dt_str)
		continue

	sheet_st = 'table'
	wb = load_workbook(path+file)
	ws = wb.get_sheet_by_name(sheet_st)
	row_len = ws.max_row+2
	dict = {}
	dict_row = {}
	day_list = []
	for index in range(2, ws.max_row+1):
		rx = row_len-index
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w2 = "%06d" % (w2)
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value
		w7 = ws.cell(row = rx, column = 7).value

		vol = 0
		if w4==u'买':
			vol = w5
		elif w4==u'卖':
			vol = 0-w5
		else:
			print "????"
			continue
		last_vol = None
		#在last_cc列
		if w2 in dict.iterkeys():
			dict[w2] = dict[w2] + vol
			day_list[dict_row[w2]][cc_idx] = None
			last_vol = dict[w2]
			dict_row[w2] = index-2
		else:
			dict[w2] = vol
			dict_row[w2] = index-2
			last_vol = vol
		temp_list = [int(dt_str),w1,w2,w3,w4,w5,w6,w7,dict[w2],last_vol]
		day_list.append(temp_list)
	if len(day_list)>0:
		df1 = pd.DataFrame(day_list, columns=c_list)
		df = df.append(df1)

if len(df)>0:
	filename = "%s%s%s%s_%s.xlsx" %(path, "trade\\", "statics_", st_date, ed_date)
	df.to_excel(filename)
